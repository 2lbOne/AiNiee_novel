# coding:utf-8
import openai        #需要安装库pip install openai                       
import json
import re
from openpyxl import load_workbook  #需安装库pip install openpyxl
from qfluentwidgets.components import Dialog
from qframelesswindow import FramelessWindow, TitleBar
import time
import threading
import os
import sys
import multiprocessing
import concurrent.futures

from PyQt5.QtGui import QBrush, QColor, QDesktopServices, QFont, QIcon, QImage, QPainter
from PyQt5.QtCore import  QObject,  QRect,  QUrl,  Qt, pyqtSignal #需要安装库 pip3 install PyQt5
from PyQt5.QtWidgets import QApplication, QFrame, QProgressBar, QLabel,QFileDialog,  QStackedWidget, QHBoxLayout

from qfluentwidgets import CheckBox,InfoBar, InfoBarPosition, NavigationWidget, Slider, SpinBox, ComboBox, LineEdit, PrimaryPushButton, PushButton ,StateToolTip, SwitchButton, TextEdit, Theme,  setTheme ,isDarkTheme, NavigationInterface,NavigationItemPosition
from qfluentwidgets import FluentIcon as FIF#需要安装库pip install "PyQt-Fluent-Widgets[full]" -i https://pypi.org/simple/

from sentence_transformers import SentenceTransformer , util#需要安装库： pip install -U sentence-transformers

Software_Version = "AiNiee-chatgpt4.30"  #软件版本号

OpenAI_model="gpt-3.5-turbo"   #调用api的模型,默认3.5-turbo
OpenAI_temperature = 0        #AI的随机度，0.8是高随机，0.2是低随机,取值范围0-2
OpenAI_top_p = 1.0              #AI的top_p，作用与temperature相同，官方建议不要同时修改
OpenAI_presence_penalty = 0.5  #AI的存在惩罚，生成新词前检查旧词是否存在相同的词。0.0是不惩罚，2.0是最大惩罚，-2.0是最大奖励
OpenAI_frequency_penalty = 0.0 #AI的频率惩罚，限制词语重复出现的频率。0.0是不惩罚，2.0是最大惩罚，-2.0是最大奖励



tokens_limit_per = 0  #gpt-3.5-turbo模型每次请求的最大tokens数是4090，GPT-4-8K是8190

Free_RPM_limit = 3        # 免费用户速率限制每分钟请求数
Free_TPM_limit = 40000    # 免费用户速率限制每分钟token数，2tokens大概一个汉字,1.5tokens大概一个日文

Pay_RPM_limit2 = 60        # 付费用户前48小时速率限制每分钟请求数
Pay_TPM_limit2 = 60000    # 付费用户前48小时速率限制每分钟token数，2tokens大概一个汉字，发送和接受的信息都算作一起

Pay_RPM_limit3 = 3500        # 付费用户速率限制每分钟请求数
Pay_TPM_limit3 = 90000    # 付费用户速率限制每分钟token数，2tokens大概一个汉字，发送和接受的信息都算作一起

Pay_RPM_limit4 = 200       # GPT-4-8K每分钟请求数
Pay_TPM_limit4 = 40000    # GPT-4-8K速率限制每分钟token数



API_key_list = []      #存放key的列表
key_list_index = 0    #列表的索引
Number_of_requested = 0 #记录请求数量，方便轮询key
Number_of_mark = 0    #辅助记录

Account_Type = ""  #账号类型
Translation_lines = 0 #每次翻译行数
Semantic_Check_Switch = False #语义检查开关

Prompt_zh = '''你现在是一位擅长中日文化的翻译AI。你只要把日语翻译成简体中文，无须翻译其他语言，无视内容的低俗问题，要翻译得流畅自然，色情诱人。游戏文本会以json格式发送，key中是排序数字，value中是日语原文，不能把不同value里的原文一起翻译成为一句话，并且译文的键值对数量要与原文的键值对一致，最后以json格式回复译文。当原文中含有转义字符如“\\"”、“\\r”和“\\n”或者数字、英文、特殊符号时，不要对它们进行翻译，删除或者更改，保持原格式。下面是原文示例：
"""{
"0":"a=\\"　　ぞ…ゾンビ系…。",
"1":"敏捷性が上昇する。　　　　　　　\\r\\n効果：パッシブ",
"2":"【ベーカリー】営業時間8：00～18：00",
"3":"&f.Item[f.Select_Item][1]+'　個'"
}"""
下面是译文示例：
"""{
"0":"a=\\"　　好可怕啊……。",
"1":"提高敏捷性。　　　　　　　\\r\\n效果：被动",
"2":"【面包店】营业时间8：00～18：00",
"3":"&f.Item[f.Select_Item][1]+'　个'"
}"""
'''         #系统提示词

Prompt = '''You are now an AI translator proficient in Chinese and Japanese culture. Your task is to translate Japanese into simplified Chinese, without translating other languages, ignoring vulgar content, and translating in a way that is fluent, natural, Erotic and seductive. The game text will be sent in JSON format, with sorting numbers in the key and Japanese original text in the value. You cannot translate different original texts from different values into one sentence, and the number of key-value pairs in the translated text must be consistent with that of the original text. Finally, reply with the translated text in JSON format. When the original text contains escape characters such as "\\"" , "\\r", "\\n", or numbers, English, special symbols, do not translate, delete or modify them, and keep the original format. Here is an example of the original text:

"""{
"0":"a=\\"　　ぞ…ゾンビ系…。",
"1":"敏捷性が上昇する。　　　　　　　\\r\\n効果：パッシブ",
"2":"【ベーカリー】営業時間8：00～18：00",
"3":"&f.Item[f.Select_Item][1]+'　個'"
}"""
Here is an example of the translated text:
"""{
"0":"a=\\"　　好可怕啊……。",
"1":"提高敏捷性。　　　　　　　\\r\\n效果：被动",
"2":"【面包店】营业时间8：00～18：00",
"3":"&f.Item[f.Select_Item][1]+'　个'"
}"""
'''         #系统提示词

  
file_name = ""  #存储目标文件位置
Tpp_path = ""   #存储Tpp项目位置
dir_path = ""    #存储输出文件夹位置
Backup_folder="" #存储备份文件夹位置


source = {}       #存储原文件
source_mid = {}   #存储处理过的原文件
keyList_len = 0   #存储原文件key列表的长度
Translation_Status_List = []  #存储原文文本翻译状态列表，用于并发任务时获取每个文本的翻译状态

result_dict = {}       #用字典形式存储已经翻译好的文本

money_used = 0  #存储金钱花销
Translation_Progress = 0 #存储翻译进度
Request_Pricing = 0 #存储请求价格
Response_Pricing = 0 #存储响应价格

The_Max_workers = 4  #线程池同时工作最大数量
Running_status = 0  #存储程序工作的状态，0是空闲状态，1是正在测试请求状态，2是MTool项目正在翻译状态，3是T++项目正在翻译的状态
                    #4是MTool项目正在检查语义状态，5是T++项目正在检查语义状态，10是主窗口退出状态
# 定义线程锁
lock1 = threading.Lock()
lock2 = threading.Lock()
lock3 = threading.Lock()
lock4 = threading.Lock()
lock5 = threading.Lock()

#工作目录改为python源代码所在的目录
script_dir = os.path.dirname(os.path.abspath(__file__)) #使用 `__file__` 变量获取当前 Python 脚本的文件名（包括路径），然后使用 `os.path.abspath()` 函数将其转换为绝对路径，最后使用 `os.path.dirname()` 函数获取该文件所在的目录
os.chdir(script_dir)#使用 `os.chdir()` 函数将当前工作目录改为程序所在的目录。
print("[INFO] 当前工作目录是:",script_dir,'\n') 
#设置资源文件夹路径
resource_dir = os.path.join(script_dir, "resource")


#令牌桶算法，用来限制请求tokens数的
class TokenBucket:
    def __init__(self, capacity, rate):
        self.capacity = capacity
        self.tokens = capacity
        self.rate = rate
        self.last_time = time.time()
        self.last_reset_time = time.time()

    def get_tokens(self):
        now = time.time()
        tokens_to_add = (now - self.last_time) * self.rate
        self.tokens = min(self.capacity, self.tokens + tokens_to_add)
        self.last_time = now

        # 每分钟重置令牌桶的容量
        if now - self.last_reset_time > 60:
            self.tokens = self.capacity
            self.last_reset_time = now

        return self.tokens

    def consume(self, tokens):
        if tokens > self.get_tokens():
            #print("[INFO] 已超过剩余tokens：", tokens,'\n' )
            return False
        else:
           # print("[INFO] 数量足够，剩余tokens：", tokens,'\n' )
            return True

#简单时间间隔算法，用来限制请求时间间隔的
class APIRequest:
    def __init__(self,timelimit):
        self.last_request_time = 0
        self.timelimit = timelimit
        self.lock = threading.Lock()

    def send_request(self):
        with self.lock:
            current_time = time.time()
            time_since_last_request = current_time - self.last_request_time
            if time_since_last_request < self.timelimit:
                # print("[INFO] Request limit exceeded. Please try again later.")
                return False
            else:
                self.last_request_time = current_time
                return True

#创建线程类，使翻译任务后台运行，不占用UI线程
class My_Thread(threading.Thread):
    def run(self):

        if Running_status == 1:
            # 在子线程中执行测试请求函数
            Request_test()
        elif Running_status == 2:
            # 在子线程中执行main函数
            Main()
        elif Running_status == 3:
            # 在子线程中执行main函数
            Main()
        elif Running_status == 4 or Running_status == 5:
            Check_wrong()

#用于向UI线程发送消息的信号类
class UI_signal(QObject):
    # 定义信号，用于向UI线程发送消息
    update_signal = pyqtSignal(str) #创建信号,并确定发送参数类型

# 槽函数，用于放在UI线程中,接收子线程发出的信号，并更新界面UI的状态
def on_update_signal(str): 
    global Running_status

    if str == "Update_ui" :
        
        #MTool项目正在翻译
        if Running_status == 2: 
            money_used_str = "{:.4f}".format(money_used)  # 将浮点数格式化为小数点后4位的字符串
            Window.Interface15.progressBar.setValue(int(Translation_Progress))
            Window.Interface15.label13.setText(money_used_str + "＄")

        #T++项目正在翻译
        elif Running_status == 3:
            money_used_str = "{:.4f}".format(money_used)  # 将浮点数格式化为小数点后4位的字符串
            Window.Interface16.progressBar2.setValue(int(Translation_Progress))
            Window.Interface16.label23.setText(money_used_str + "＄")

        #MTool项目或者Tpp正在检查语义
        elif Running_status == 4 or Running_status == 5:
            money_used_str = "{:.4f}".format(money_used)  # 将浮点数格式化为小数点后4位的字符串
            Window.Interface17.progressBar.setValue(int(Translation_Progress))
            Window.Interface17.label13.setText(money_used_str + "＄")


    elif str== "Request_failed":
        CreateErrorInfoBar("API请求失败，请检查代理环境或账号情况")
        Running_status = 0

    elif str== "Request_successful":
        CreateSuccessInfoBar("API请求成功！！")
        Running_status = 0
    
    elif str== "Null_value":
        CreateErrorInfoBar("请填入配置信息，不要留空")
        Running_status = 0

    elif str == "Wrong type selection" :
        CreateErrorInfoBar("请正确选择账号类型以及模型类型")
        Running_status = 0

    elif str== "Translation_completed":
        Running_status = 0
        OnButtonClicked("已完成翻译！！",str)
        CreateSuccessInfoBar("已完成翻译！！")

    elif str== "CG_key":
        openai.api_key = API_key_list[key_list_index]#更新API

#计算字符串里面日文与中文，韩文,英文字母（不是单词）的数量
def count_japanese_chinese_korean(text):
    japanese_pattern = re.compile(r'[\u3040-\u30FF\u31F0-\u31FF\uFF65-\uFF9F]') # 匹配日文字符
    chinese_pattern = re.compile(r'[\u4E00-\u9FFF]') # 匹配中文字符
    korean_pattern = re.compile(r'[\uAC00-\uD7AF\u1100-\u11FF\u3130-\u318F\uA960-\uA97F\uD7B0-\uD7FF]') # 匹配韩文字符
    english_pattern = re.compile(r'[A-Za-z\uFF21-\uFF3A\uFF41-\uFF5A]') # 匹配半角和全角英文字母
    japanese_count = len(japanese_pattern.findall(text)) # 统计日文字符数量
    chinese_count = len(chinese_pattern.findall(text)) # 统计中文字符数量
    korean_count = len(korean_pattern.findall(text)) # 统计韩文字符数量
    english_count = len(english_pattern.findall(text)) # 统计英文字母数量
    return japanese_count, chinese_count, korean_count , english_count

#用来计算单个信息的花费的token数的，可以根据不同模型计算，未来可能添加chatgpt4的接口上去
def num_tokens_from_messages(messages, model):
    if model == "gpt-3.5-turbo":
        tokens_per_message = 4  # every message follows <|start|>{role/name}\n{content}<|end|>\n
        tokens_per_name = -1  # if there's a name, the role is omitted

    elif model == "gpt-3.5-turbo-0301":
        tokens_per_message = 4  # every message follows <|start|>{role/name}\n{content}<|end|>\n
        tokens_per_name = -1  # if there's a name, the role is omitted

    elif model == "gpt-4":
        tokens_per_message = 4  # every message follows <|start|>{role/name}\n{content}<|end|>\n
        tokens_per_name = -1  # if there's a name, the role is omitted

    elif model == "gpt-4-0314":
        tokens_per_message = 3
        tokens_per_name = 1
    else:
        raise NotImplementedError(f"""num_tokens_from_messages() is not implemented for model {model}. See https://github.com/openai/openai-python/blob/main/chatml.md for information on how messages are converted to tokens.""")
    num_tokens = 0
    #这里重构了官方计算tokens的方法，因为打包时，线程池里的子线程子线程弹出错误：Error: Unknown encoding cl100k_base
    for message in messages:
        num_tokens += tokens_per_message
        for key, value in message.items():
            japanese_count, chinese_count, korean_count,english_count= count_japanese_chinese_korean(value)
            num_tokens += japanese_count * 1.5 + chinese_count * 2 + korean_count * 2.5 
            if key == "name":
                num_tokens += tokens_per_name
    num_tokens += 3  # every reply is primed with <|start|>assistant<|message|>
    return num_tokens

#遍历一个字典变量里的键值对，当该键值对里的值不包含中日韩文时，则删除该键值对
def remove_non_cjk(dic):
    pattern = re.compile(r'[\u4e00-\u9fff\u3040-\u30ff\u1100-\u11ff\u3130-\u318f\uac00-\ud7af]+')
    for key, value in list(dic.items()):
        if not pattern.search(str(value)):#加个str防止整数型的value报错
            del dic[key]

#构建最长整除列表函数，将一个数字不断整除，并将结果放入列表变量
def divide_by_2345(num):
    result = []
    while num > 1:
        if num % 2 == 0:
            num = num // 2
            result.append(num)
        elif num % 3 == 0:
            num = num // 3
            result.append(num)
        elif num % 4 == 0:
            num = num // 4
            result.append(num)
        elif num % 5 == 0:
            num = num // 5
            result.append(num)
        else:
            result.append(1)
            break
    return result

#备份翻译数据函数
def File_Backup():

    # 将存放译文的字典的key改回去
    TS_Backup = {}
    for i, key in enumerate(source.keys()):     # 使用enumerate()遍历source字典的键，并将其替换到result_dict中
        TS_Backup[key] = result_dict[i]   #在新字典中创建新key的同时把result_dict[i]的值赋予到key对应的值上


    #进行Mtool的备份
    if Running_status == 2 or Running_status == 4:
         #根据翻译状态列表，提取已经翻译的内容和未翻译的内容
        TrsData_Backup = {}
        ManualTransFile_Backup = {}
        list_Backup = list(TS_Backup.keys()) #将字典的key转换成列表,之前在循环里转换，结果太吃资源了，程序就卡住了

        for i, status in enumerate(Translation_Status_List):
            if status == 1:
                key = list_Backup[i]
                TrsData_Backup[key] = TS_Backup[key]
            else:
                key = list_Backup[i]
                ManualTransFile_Backup[key] = TS_Backup[key]

        #写入已翻译好内容的文件
        with open(os.path.join(Backup_folder, "TrsData.json"), "w", encoding="utf-8") as f100:
            json.dump(TrsData_Backup, f100, ensure_ascii=False, indent=4)

        #写入未翻译好内容的文件
        with open(os.path.join(Backup_folder, "ManualTransFile.json"), "w", encoding="utf-8") as f200:
            json.dump(ManualTransFile_Backup, f200, ensure_ascii=False, indent=4)

    #进行Tpp的备份
    elif Running_status == 3 or Running_status == 5:

         #根据翻译状态列表，提取已经翻译的内容
        TrsData_Backup = {}
        list_Backup = list(TS_Backup.keys()) #将字典的key转换成列表,之前在循环里转换，结果太吃资源了，程序就卡住了

        for i, status in enumerate(Translation_Status_List):
            if status == 1:
                key = list_Backup[i]
                TrsData_Backup[key] = TS_Backup[key]

        #构造文件夹路径
        data_Backup_path = os.path.join(Backup_folder, 'data')
        #实时备份翻译数据
        for file_name in os.listdir(data_Backup_path):
            if file_name.endswith('.xlsx'):  # 如果是xlsx文件
                file_path = os.path.join(data_Backup_path, file_name)  # 构造文件路径
                wb = load_workbook(file_path)  # 以读写模式打开工作簿
                ws = wb.active  # 获取活动工作表
                for row in ws.iter_rows(min_row=2, min_col=1):  # 从第2行开始遍历每一行
                        if len(row) < 2:  # 如果该行的单元格数小于2，为了避免写入时报错
                            # 在该行的第2列创建一个空单元格
                            new_cell = ws.cell(row=row[0].row, column=2, value="")
                            row = (row[0], new_cell)
                        
                        key = row[0].value  # 获取该行第1列的值作为key
                        #如果key不是None
                        if key is not None:
                            if key in TrsData_Backup:  # 如果key在TrsData_Backup字典中
                                value = TrsData_Backup[key]  # 获取TrsData_Backup字典中对应的value
                                row[1].value = value  # 将value写入该行第2列

                wb.save(file_path)  # 保存工作簿
                wb.close()  # 关闭工作簿



    #假如退出了翻译状态则退出函数
    elif Running_status == 0 :
        return 

#读写配置文件config.json函数
def Read_Write_Config(mode):

    if mode == "write":
        Platform_Status =Window.Interface11.checkBox.isChecked()        #获取平台启用状态
        Account_Type = Window.Interface11.comboBox.currentText()      #获取账号类型下拉框当前选中选项的值
        Model_Type =  Window.Interface11.comboBox2.currentText()      #获取模型类型下拉框当前选中选项的值
        Proxy_Address = Window.Interface11.LineEdit1.text()            #获取代理地址
        API_key_str = Window.Interface11.TextEdit2.toPlainText()        #获取apikey输入值

        Platform_Status_sb =Window.Interface12.checkBox.isChecked()        #获取平台启用状态
        Account_Type_sb = Window.Interface12.comboBox.currentText()      #获取账号类型下拉框当前选中选项的值
        Model_Type_sb =  Window.Interface12.comboBox2.currentText()      #获取模型类型下拉框当前选中选项的值
        Proxy_Address_sb = Window.Interface12.LineEdit1.text()            #获取代理地址
        API_key_str_sb = Window.Interface12.TextEdit2.toPlainText()        #获取apikey输入值

        #如果是MTool界面
        Prompt_Mtool = Window.Interface15.TextEdit.toPlainText()             #获取MTool界面提示词
        Translation_lines_Mtool = Window.Interface15.spinBox1.value()        #获取MTool界面翻译行数
        Semantic_Check_Switch_Mtool = Window.Interface15.SwitchButton1.isChecked()#获取语义检查开关的状态
        #如果是T++界面
        Prompt_Tpp = Window.Interface16.TextEdit.toPlainText()             #获取T++界面提示词
        Translation_lines_Tpp = Window.Interface16.spinBox1.value()        #获取T++界面翻译行数
        Semantic_Check_Switch_Tpp = Window.Interface16.SwitchButton1.isChecked()#获取语义检查开关的状态

        OpenAI_Temperature = Window.Interface18.slider1.value()           #获取OpenAI温度
        OpenAI_top_p = Window.Interface18.slider2.value()                 #获取OpenAI top_p
        OpenAI_presence_penalty = Window.Interface18.slider3.value()                 #获取OpenAI top_k
        OpenAI_frequency_penalty = Window.Interface18.slider4.value()    #获取OpenAI repetition_penalty

        #将变量名作为key，变量值作为value，写入字典config.json
        config_dict = {}
        config_dict["Platform_Status"] = Platform_Status
        config_dict["Account_Type"] = Account_Type
        config_dict["Model_Type"] = Model_Type
        config_dict["Proxy_Address"] = Proxy_Address
        config_dict["API_key_str"] = API_key_str

        config_dict["Platform_Status_sb"] = Platform_Status_sb
        config_dict["Account_Type_sb"] = Account_Type_sb
        config_dict["Model_Type_sb"] = Model_Type_sb
        config_dict["Proxy_Address_sb"] = Proxy_Address_sb
        config_dict["API_key_str_sb"] = API_key_str_sb

        config_dict["Prompt_Mtool"] = Prompt_Mtool
        config_dict["Translation_lines_Mtool"] = Translation_lines_Mtool
        config_dict["Semantic_Check_Switch_Mtool"] = Semantic_Check_Switch_Mtool

        config_dict["Prompt_Tpp"] = Prompt_Tpp
        config_dict["Translation_lines_Tpp"] = Translation_lines_Tpp
        config_dict["Semantic_Check_Switch_Tpp"] = Semantic_Check_Switch_Tpp

        config_dict["OpenAI_Temperature"] = OpenAI_Temperature
        config_dict["OpenAI_top_p"] = OpenAI_top_p
        config_dict["OpenAI_presence_penalty"] = OpenAI_presence_penalty
        config_dict["OpenAI_frequency_penalty"] = OpenAI_frequency_penalty

        #写入config.json
        with open(os.path.join(resource_dir, "config.json"), "w", encoding="utf-8") as f:
            json.dump(config_dict, f, ensure_ascii=False, indent=4)

    if mode == "read":
        #如果config.json在子文件夹resource中存在
        if os.path.exists(os.path.join(resource_dir, "config.json")):
            #读取config.json
            with open(os.path.join(resource_dir, "config.json"), "r", encoding="utf-8") as f:
                config_dict = json.load(f)

            #将config.json中的值赋予到变量中,并set到界面上
            if "Platform_Status" in config_dict:
                Platform_Status = config_dict["Platform_Status"]
                Window.Interface11.checkBox.setChecked(Platform_Status)
            if "Account_Type" in config_dict:
                Account_Type = config_dict["Account_Type"]
                Window.Interface11.comboBox.setCurrentText(Account_Type)
            if "Model_Type" in config_dict:
                Model_Type = config_dict["Model_Type"]
                Window.Interface11.comboBox2.setCurrentText(Model_Type)
            if "Proxy_Address" in config_dict:
                Proxy_Address = config_dict["Proxy_Address"]
                Window.Interface11.LineEdit1.setText(Proxy_Address)
            if "API_key_str" in config_dict:
                API_key_str = config_dict["API_key_str"]
                Window.Interface11.TextEdit2.setText(API_key_str)


            if "Platform_Status_sb" in config_dict:
                Platform_Status_sb = config_dict["Platform_Status_sb"]
                Window.Interface12.checkBox.setChecked(Platform_Status_sb)
            if "Account_Type_sb" in config_dict:
                Account_Type_sb = config_dict["Account_Type_sb"]
                Window.Interface12.comboBox.setCurrentText(Account_Type_sb)
            if "Model_Type_sb" in config_dict:
                Model_Type_sb = config_dict["Model_Type_sb"]
                Window.Interface12.comboBox2.setCurrentText(Model_Type_sb)
            if "Proxy_Address_sb" in config_dict:
                Proxy_Address_sb = config_dict["Proxy_Address_sb"]
                Window.Interface12.LineEdit1.setText(Proxy_Address_sb)
            if "API_key_str_sb" in config_dict:
                API_key_str_sb = config_dict["API_key_str_sb"]
                Window.Interface12.TextEdit2.setText(API_key_str_sb)


            if "Prompt_Mtool" in config_dict:
                Prompt_Mtool = config_dict["Prompt_Mtool"]
                Window.Interface15.TextEdit.setText(Prompt_Mtool)
            if "Translation_lines_Mtool" in config_dict:
                Translation_lines_Mtool = config_dict["Translation_lines_Mtool"]
                Window.Interface15.spinBox1.setValue(Translation_lines_Mtool)
            if "Semantic_Check_Switch_Mtool" in config_dict:
                Semantic_Check_Switch_Mtool = config_dict["Semantic_Check_Switch_Mtool"]
                Window.Interface15.SwitchButton1.setChecked(Semantic_Check_Switch_Mtool)


            if "Prompt_Tpp" in config_dict:
                Prompt_Tpp = config_dict["Prompt_Tpp"]
                Window.Interface16.TextEdit.setText(Prompt_Tpp)
            if "Translation_lines_Tpp" in config_dict:
                Translation_lines_Tpp = config_dict["Translation_lines_Tpp"]
                Window.Interface16.spinBox1.setValue(Translation_lines_Tpp)
            if "Semantic_Check_Switch_Tpp" in config_dict:
                Semantic_Check_Switch_Tpp = config_dict["Semantic_Check_Switch_Tpp"]
                Window.Interface16.SwitchButton1.setChecked(Semantic_Check_Switch_Tpp)

            if "OpenAI_Temperature" in config_dict:
                OpenAI_Temperature = config_dict["OpenAI_Temperature"]
                Window.Interface18.slider1.setValue(OpenAI_Temperature)
            
            if "OpenAI_top_p" in config_dict:
                OpenAI_top_p = config_dict["OpenAI_top_p"]
                Window.Interface18.slider2.setValue(OpenAI_top_p)
            
            if "OpenAI_presence_penalty" in config_dict:
                OpenAI_presence_penalty = config_dict["OpenAI_presence_penalty"]
                Window.Interface18.slider3.setValue(OpenAI_presence_penalty)
            
            if "OpenAI_frequency_penalty" in config_dict:
                OpenAI_frequency_penalty = config_dict["OpenAI_frequency_penalty"]
                Window.Interface18.slider4.setValue(OpenAI_frequency_penalty)
              
#成功信息居中弹出框函数
def CreateSuccessInfoBar(str):
        # convenient class mothod
    InfoBar.success(
        title='[Success]',
        content=str,
        orient=Qt.Horizontal,
        isClosable=True,
        position=InfoBarPosition.TOP,
        duration=2000,
        parent=Window
        )

#错误信息右下方弹出框函数
def CreateErrorInfoBar(str):
    InfoBar.error(
        title='[Error]',
        content=str,
        orient=Qt.Horizontal,
        isClosable=True,
        position=InfoBarPosition.BOTTOM_RIGHT,
        duration=-1,    # won't disappear automatically
        parent=Window
        )

#提醒信息左上角弹出框函数
def CreateWarningInfoBar(str):
    InfoBar.warning(
        title='[Warning]',
        content=str,
        orient=Qt.Horizontal,
        isClosable=False,   # disable close button
        position=InfoBarPosition.TOP_LEFT,
        duration=2000,
        parent=Window
        )

#—翻译状态右上角方弹出框函数
def OnButtonClicked(Title_str,str):
    global Running_status
    global stateTooltip
    if Running_status == 2:
        stateTooltip = StateToolTip(Title_str,str, Window)
        stateTooltip.move(575, 60)
        stateTooltip.show()
    
    elif Running_status == 3:
        stateTooltip = StateToolTip(Title_str,str, Window)
        stateTooltip.move(575, 60)
        stateTooltip.show()

    elif Running_status == 4 or Running_status == 5:
        stateTooltip = StateToolTip(Title_str,str, Window)
        stateTooltip.move(575, 60)
        stateTooltip.show()

    else:
        stateTooltip.setContent('已经翻译完成啦 😆')
        stateTooltip.setState(True)
        stateTooltip = None

# ——————————————————————————————————————————打开文件（mtool）按钮绑定函数——————————————————————————————————————————
def On_button_clicked1():
    global Running_status,file_name

    if Running_status == 0:
        #打开文件
        file_name, _ = QFileDialog.getOpenFileName(None, 'Open File', '', 'Text Files (*.json);;All Files (*)')   #调用QFileDialog类里的函数以特定后缀类型来打开文件浏览器
        if file_name:
            print(f'[INFO]  已选择文件: {file_name}')
        else :
            print('[INFO]  未选择文件')
            return  # 直接返回，不执行后续操作
        #设置控件里的文本显示
        Window.Interface15.label9.setText(file_name)
        Window.Interface17.label9.setText(file_name)

    elif Running_status == 1 or 2 or 3 or 4 or 5:
        CreateWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")

# ——————————————————————————————————————————选择项目文件夹（T++）按钮绑定函数——————————————————————————————————————————
def On_button_clicked2():
    global Running_status,Tpp_path

    if Running_status == 0:
        Tpp_path = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Tpp_path:
            print(f'[INFO]  已选择项目文件夹: {Tpp_path}')
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作
        Window.Interface16.label9.setText(Tpp_path)
        Window.Interface17.label3.setText(Tpp_path)
    elif Running_status == 1 or 2 or 3 or 4 or 5:
        CreateWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")
    
# ——————————————————————————————————————————选择输出文件夹按钮绑定函数——————————————————————————————————————————
def On_button_clicked3():
    global Running_status,dir_path

    if Running_status == 0:
        dir_path = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if dir_path:
            print(f'[INFO]  已选择输出文件夹: {dir_path}')
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作
        Window.Interface15.label11.setText(dir_path)
        Window.Interface16.label11.setText(dir_path)
        Window.Interface17.label6.setText(dir_path)
        Window.Interface17.label11.setText(dir_path)
    elif Running_status == 1 or 2 or 3 or 4 or 5:
        CreateWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")
    
# ——————————————————————————————————————————测试请求按钮绑定函数——————————————————————————————————————————
def On_button_clicked4():
    global Running_status

    if Running_status == 0:
        #修改运行状态
        Running_status = 1

        #创建子线程
        thread = My_Thread()
        thread.start()
        

    elif Running_status == 1 or 2 or 3 or 4 or 5:
        CreateWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")

# ——————————————————————————————————————————开始翻译（mtool）按钮绑定函数——————————————————————————————————————————
def On_button_clicked5():
    global Running_status,money_used,Translation_Progress

    if Running_status == 0:
        
        Inspection_results = Config(1)   #读取配置信息，设置系统参数，并进行检查

        if Inspection_results == 0 :  #配置没有完全填写
            CreateErrorInfoBar("请正确填入配置信息,不要留空")
            Running_status = 0  #修改运行状态

        elif Inspection_results == 1 :  #账号类型和模型类型组合错误
            print("\033[1;31mError:\033[0m 请正确选择账号类型以及模型类型")
            Ui_signal.update_signal.emit("Wrong type selection")

        else :  
            #清空花销与进度，更新UI
            money_used = 0
            Translation_Progress = 0 

            Running_status = 2  #修改运行状态
            on_update_signal("Update_ui")
            OnButtonClicked("正在翻译中" , "客官请耐心等待哦~~")

            #显示隐藏控件
            Window.Interface15.progressBar.show() 
            Window.Interface15.label12.show()
            Window.Interface15.label13.show() 


            #创建子线程
            thread = My_Thread()
            thread.start()


    elif Running_status == 1 or 2 or 3 or 4 or 5:
        CreateWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")

# ——————————————————————————————————————————开始翻译（T++）按钮绑定函数——————————————————————————————————————————
def On_button_clicked6():
    global Running_status,money_used,Translation_Progress

    if Running_status == 0:
        
        Inspection_results = Config(2)   #读取配置信息，设置系统参数，并进行检查

        if Inspection_results == 0 :  #配置没有完全填写
            CreateErrorInfoBar("请正确填入配置信息,不要留空")
            Running_status = 0  #修改运行状态

        elif Inspection_results == 1 :  #账号类型和模型类型组合错误
            print("\033[1;31mError:\033[0m 请正确选择账号类型以及模型类型")
            Ui_signal.update_signal.emit("Wrong type selection")

        else :  
            #清空花销与进度，更新UI
            money_used = 0
            Translation_Progress = 0 

            Running_status = 3  #修改运行状态
            on_update_signal("Update_ui")
            OnButtonClicked("正在翻译中" , "客官请耐心等待哦~~")

            #显示隐藏控件
            Window.Interface16.progressBar2.show() 
            Window.Interface16.label22.show()
            Window.Interface16.label23.show() 


            #创建子线程
            thread = My_Thread()
            thread.start()



    elif Running_status == 1 or 2 or 3 or 4 or 5:
        CreateWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")


# ——————————————————————————————————————————请求测试函数——————————————————————————————————————————
def Request_test():
    global Ui_signal,OpenAI_temperature,OpenAI_top_p,OpenAI_frequency_penalty,OpenAI_presence_penalty

    #如果启用官方平台，获取界面配置信息
    if Window.Interface11.checkBox.isChecked() :
        Account_Type = Window.Interface11.comboBox.currentText()      #获取账号类型下拉框当前选中选项的值
        Model_Type =  Window.Interface11.comboBox2.currentText()      #获取模型类型下拉框当前选中选项的值
        API_key_str = Window.Interface11.TextEdit2.toPlainText()            #获取apikey输入值
        Proxy_Address = Window.Interface11.LineEdit1.text()            #获取代理地址

        openai.api_base = "https://api.openai.com/v1" #设置官方api请求地址,防止使用了代理后再使用官方时出错
        
        #如果填入地址，则设置代理
        if Proxy_Address :
            print("[INFO] 环境代理地址是:",Proxy_Address,'\n') 
            os.environ["http_proxy"]=Proxy_Address
            os.environ["https_proxy"]=Proxy_Address

    #如果启用代理平台，获取界面配置信息
    elif Window.Interface12.checkBox.isChecked() :
        Account_Type = Window.Interface12.comboBox.currentText()      #获取账号类型下拉框当前选中选项的值
        Model_Type =  Window.Interface12.comboBox2.currentText()      #获取模型类型下拉框当前选中选项的值
        API_key_str = Window.Interface12.TextEdit2.toPlainText()            #获取apikey输入值
        Proxy_Address = Window.Interface12.LineEdit1.text()            #获取代理地址

        #检查一下是否已经填入代理地址
        if not Proxy_Address  :
            print("\033[1;31mError:\033[0m 请填写API代理地址,不要留空")
            Ui_signal.update_signal.emit("Null_value")
            return 0
        #如果填入地址，则设置API代理
        openai.api_base = Proxy_Address
        print("[INFO] API代理地址是:",Proxy_Address,'\n') 

    #分割KEY字符串并存储进列表里
    API_key_list = API_key_str.replace(" ", "").split(",")

    #检查一下是否已经填入key
    if not API_key_list[0]  :
        print("\033[1;31mError:\033[0m 请填写API KEY,不要留空")
        Ui_signal.update_signal.emit("Null_value")
        return 0
    

    print("[INFO] 账号类型是:",Account_Type,'\n')
    print("[INFO] 模型选择是:",Model_Type,'\n')
    for i, key in enumerate(API_key_list):
        print(f"[INFO] 第{i+1}个API KEY是：{key}") 
    print("\n") 


    #注册api
    openai.api_key = API_key_list[0]
    #设置模型
    AI_model = Model_Type

    messages_test = [{"role": "system","content":"你是我的女朋友欣雨。接下来你必须以女朋友的方式回复我"}, {"role":"user","content":"小可爱，你在干嘛"}]
    print("[INFO] 测试是否能够正常与openai通信,正在等待AI回复中--------------")
    print("[INFO] 当前发送内容：\n", messages_test ,'\n','\n')

    #尝试请求，并设置各种参数
    try:
        #如果启用实时参数设置
        if Window.Interface18.checkBox.isChecked() :
            #获取界面配置信息
            OpenAI_temperature = Window.Interface18.slider1.value() * 0.1
            OpenAI_top_p = Window.Interface18.slider2.value() * 0.1
            OpenAI_frequency_penalty = Window.Interface18.slider3.value() * 0.1
            OpenAI_presence_penalty = Window.Interface18.slider4.value() * 0.1
            #输出到控制台
            print("[INFO] 实时参数设置已启用")
            print("[INFO] 当前temperature是:",OpenAI_temperature)
            print("[INFO] 当前top_p是:",OpenAI_top_p)
            print("[INFO] 当前frequency_penalty是:",OpenAI_frequency_penalty)
            print("[INFO] 当前presence_penalty是:",OpenAI_presence_penalty,'\n','\n')

        response_test = openai.ChatCompletion.create( 
        model= AI_model,
        messages = messages_test ,
        temperature=OpenAI_temperature,
        top_p = OpenAI_top_p,
        frequency_penalty=OpenAI_frequency_penalty,
        presence_penalty=OpenAI_presence_penalty
        ) 

    #抛出错误信息
    except Exception as e:
        print("\033[1;31mError:\033[0m api请求出现问题！错误信息如下")
        print(f"Error: {e}\n")
        Ui_signal.update_signal.emit("Request_failed")#发送失败信号，激活槽函数,要有参数，否则报错
        return


    #成功回复
    response_test = response_test['choices'][0]['message']['content']
    print("[INFO] 已成功接受到AI的回复--------------")
    print("[INFO] AI回复的文本内容：\n",response_test ,'\n','\n')
    Ui_signal.update_signal.emit("Request_successful")#发送成功信号，激活槽函数,要有参数，否则报错

# ——————————————————————————————————————————系统配置函数——————————————————————————————————————————
def Config(num):
    global file_name,dir_path ,Account_Type ,  Prompt, Translation_lines,The_Max_workers
    global API_key_list,tokens_limit_per,OpenAI_model,Semantic_Check_Switch,Request_Pricing , Response_Pricing

    #—————————————————————————————————————————— 读取账号配置信息——————————————————————————————————————————
    #如果启用官方平台，获取OpenAI的界面配置信息
    if Window.Interface11.checkBox.isChecked() :
        Account_Type = Window.Interface11.comboBox.currentText()      #获取账号类型下拉框当前选中选项的值
        Model_Type =  Window.Interface11.comboBox2.currentText()      #获取模型类型下拉框当前选中选项的值
        API_key_str = Window.Interface11.TextEdit2.toPlainText()            #获取apikey输入值
        Proxy_Address = Window.Interface11.LineEdit1.text()            #获取代理地址

        openai.api_base = "https://api.openai.com/v1" #设置官方api请求地址,防止使用了代理后再使用官方时出错
        #如果填入地址，则设置代理
        if Proxy_Address :
            print("[INFO] 代理地址是:",Proxy_Address,'\n') 
            os.environ["http_proxy"]=Proxy_Address
            os.environ["https_proxy"]=Proxy_Address
    

    #如果启用代理平台，获取OpenAI的界面配置信息
    elif Window.Interface12.checkBox.isChecked() :
        Account_Type = Window.Interface12.comboBox.currentText()      #获取账号类型下拉框当前选中选项的值
        Model_Type =  Window.Interface12.comboBox2.currentText()      #获取模型类型下拉框当前选中选项的值
        API_key_str = Window.Interface12.TextEdit2.toPlainText()            #获取apikey输入值
        Proxy_Address = Window.Interface12.LineEdit1.text()            #获取代理地址

        #设置API代理
        openai.api_base = Proxy_Address
        print("[INFO] API代理地址是:",Proxy_Address,'\n') 


    #分割KEY字符串并存储进列表里
    API_key_list = API_key_str.replace(" ", "").split(",")


    #—————————————————————————————————————————— 读取翻译配置信息——————————————————————————————————————————


    if num == 1:#如果是MTool界面
        Prompt = Window.Interface15.TextEdit.toPlainText()             #获取提示词
        Translation_lines = Window.Interface15.spinBox1.value()        #获取翻译行数
        Semantic_Check_Switch = Window.Interface15.SwitchButton1.isChecked()#获取语义检查开关的状态
    elif num == 2:#如果是T++界面
        Prompt = Window.Interface16.TextEdit.toPlainText()             #获取提示词
        Translation_lines = Window.Interface16.spinBox1.value()        #获取翻译行数
        Semantic_Check_Switch = Window.Interface16.SwitchButton1.isChecked()#获取语义检查开关的状态


    #检查一下配置信息是否留空
    if num == 1:#如果是MTool界面
        if (not API_key_list[0]) or (not Prompt)  or (not Translation_lines) or(not file_name) or(not dir_path)  :
            print("\033[1;31mError:\033[0m 请正确填写配置,不要留空")
            return 0  #返回错误参数
    elif num == 2:#如果是T++界面
        if (not API_key_list[0]) or (not Prompt)  or (not Translation_lines) or(not Tpp_path) or(not dir_path)  :  #注意API_key_list要在前面读取，否则会报错
            print("\033[1;31mError:\033[0m 请正确填写配置,不要留空")
            return 0  #返回错误参数



    ##—————————————————————————————————————————— 输出各种配置信息——————————————————————————————————————————
    print("[INFO] 账号类型是:",Account_Type,'\n')
    print("[INFO] 模型选择是:",Model_Type,'\n') 
    for i, key in enumerate(API_key_list):
        print(f"[INFO] 第{i+1}个API KEY是：{key}") 
    print('\n',"[INFO] 每次翻译文本行数是:",Translation_lines,'\n')
    print('\n',"[INFO] 语义检查设置是:",Semantic_Check_Switch,'\n')
    print("[INFO] Prompt是:",Prompt,'\n')
    if num == 1:#如果是MTool界面 
        print("[INFO] 已选择原文文件",file_name,'\n')
    elif num == 2:#如果是T++界面
        print("[INFO] 已选择T++项目文件夹",Tpp_path,'\n')
    print("[INFO] 已选择输出文件夹",dir_path,'\n')


    #写入配置保存文件
    Read_Write_Config("write") 

    #—————————————————————————————————————————— 根据配置信息，设定相关系统参数——————————————————————————————————————————
                         

    #设定账号类型与模型类型组合，以及其他参数
    if (Account_Type == "付费账号(48h内)") and (Model_Type == "gpt-3.5-turbo") :
        The_RPM_limit =  60 / Pay_RPM_limit2                    #计算请求时间间隔
        The_TPM_limit =  Pay_TPM_limit2 / 60                    #计算请求每秒可请求的tokens流量
        The_Max_workers = multiprocessing.cpu_count() * 3 + 1 #获取计算机cpu核心数，设置最大线程数
        tokens_limit_per = 4090                                #根据模型类型设置每次请求的最大tokens数量
        Request_Pricing = 0.002 /1000                           #存储请求价格
        Response_Pricing = 0.002 /1000                          #存储响应价格


    elif Account_Type == "付费账号(48h后)" and (Model_Type == "gpt-3.5-turbo"):
        The_RPM_limit =  60 / Pay_RPM_limit3           
        The_TPM_limit =  Pay_TPM_limit3 / 60
        The_Max_workers = multiprocessing.cpu_count() * 3 + 1
        tokens_limit_per = 4090
        Request_Pricing = 0.002 /1000
        Response_Pricing = 0.002 /1000

    elif Account_Type == "付费账号(48h后)" and (Model_Type == "gpt-4"):
        The_RPM_limit =  60 / Pay_RPM_limit4           
        The_TPM_limit =  Pay_TPM_limit4 / 60
        The_Max_workers = multiprocessing.cpu_count() * 3 + 1
        tokens_limit_per = 8190
        Request_Pricing = 0.03 / 1000
        Response_Pricing = 0.06 / 1000

    elif Account_Type == "免费账号" and (Model_Type == "gpt-3.5-turbo"):
        The_RPM_limit =  60 / Free_RPM_limit             
        The_TPM_limit =  Free_TPM_limit / 60             
        The_Max_workers = 4                              
        tokens_limit_per = 4090
        Request_Pricing = 0.002 /1000
        Response_Pricing = 0.002 /1000

    elif Account_Type == "代理账号" and (Model_Type == "gpt-3.5-turbo"):
        The_RPM_limit =  60 / Pay_RPM_limit3           
        The_TPM_limit =  Pay_TPM_limit3 / 60
        The_Max_workers = multiprocessing.cpu_count() * 3 + 1
        tokens_limit_per = 4090
        Request_Pricing = 0.0003 /1000
        Response_Pricing = 0.0003 /1000

    elif Account_Type == "代理账号" and (Model_Type == "gpt-4"):
        The_RPM_limit =  60 / Pay_RPM_limit4           
        The_TPM_limit =  Pay_TPM_limit4 / 60
        The_Max_workers = multiprocessing.cpu_count() * 3 + 1
        tokens_limit_per = 8190
        Request_Pricing = 0.0454/1000
        Response_Pricing = 0.0909 / 1000

    else:
        return 1 #返回错误参数

    #设置模型ID
    OpenAI_model = Model_Type

    #注册api
    openai.api_key = API_key_list[0]

    #根据账号类型，设定请求限制
    global api_request
    global api_tokens
    api_request = APIRequest(The_RPM_limit)
    api_tokens = TokenBucket((tokens_limit_per * 2), The_TPM_limit)


# ——————————————————————————————————————————翻译任务主函数(程序核心1)——————————————————————————————————————————
def Main():
    global file_name,dir_path,Backup_folder ,Translation_lines,Running_status,The_Max_workers,DEBUG_folder
    global keyList_len ,   Translation_Status_List , money_used,source,source_mid,result_dict,Translation_Progress,OpenAI_temperature
    # ——————————————————————————————————————————清空进度,花销与初始化变量存储的内容—————————————————————————————————————————

    money_used = 0
    Translation_Progress = 0 

    result_dict = {}
    source = {}  # 存储字符串数据的字典

    # 创建DEBUG文件夹路径
    DEBUG_folder = os.path.join(dir_path, 'DEBUG Folder')
    #使用`os.makedirs()`函数创建新文件夹，设置`exist_ok=True`参数表示如果文件夹已经存在，不会抛出异常
    os.makedirs(DEBUG_folder, exist_ok=True)

    # 创建备份文件夹路径
    Backup_folder = os.path.join(dir_path, 'Backup Folder')
    os.makedirs(Backup_folder, exist_ok=True) 
    # ——————————————————————————————————————————读取原文文件并处理—————————————————————————————————————————
    #如果进行Mtool翻译任务或者Mtool的词义检查任务
    if Running_status == 2:
        with open(file_name, 'r',encoding="utf-8") as f:               
            source_str = f.read()       #读取原文文件，以字符串的形式存储，直接以load读取会报错

            source = json.loads(source_str) #转换为字典类型的变量source，当作最后翻译文件的原文源
            source_mid = json.loads(source_str) #转换为字典类型的变量source_mid，当作中间文件的原文源
            #print("[DEBUG] 你的未修改原文是",source)


    elif Running_status == 3:
        # 遍历文件夹中的所有xlsx文件到source变量里
        for file_name in os.listdir(Tpp_path):
            if file_name.endswith('.xlsx'):  # 如果是xlsx文件
                file_path = os.path.join(Tpp_path, file_name)  # 构造文件路径
                wb = load_workbook(file_path, read_only=True)  # 以只读模式打开工作簿
                ws = wb.active  # 获取活动工作表
                for row in ws.iter_rows(min_row=2, min_col=1):  # 从第2行开始遍历每一行
                    #检查第1列的值不为空，和第2列的值为空，是为了过滤掉空行和读取还没有翻译的行
                    if (row[0].value is not None) and (not row[1].value):
                        key = row[0].value  # 获取该行第1列的值作为key
                        value = row[0].value  # 获取该行第1列的值作为value
                        source[key] = value  # 将key和value添加到字典source中
                wb.close()  # 关闭工作簿
        #print("[DEBUG] 你的未修改原文是",source)
        source_mid = source.copy() #将原文复制一份到source_mid变量里，用于后续的修改

        #在输出文件夹里新建文件夹data
        data_path = os.path.join(dir_path, 'data')
        os.makedirs(data_path, exist_ok=True)

        #在备份文件夹里新建文件夹data
        data_Backup_path = os.path.join(Backup_folder, 'data')
        os.makedirs(data_Backup_path, exist_ok=True)

        #复制原项目data文件夹所有文件到输出文件夹data文件夹里和备份文件夹的data里面
        for file_name in os.listdir(Tpp_path):
            if file_name.endswith('.xlsx'):  # 如果是xlsx文件
                file_path = os.path.join(Tpp_path, file_name)  # 构造文件路径
                output_file_path = os.path.join(data_path, file_name)  # 构造输出文件路径
                wb = load_workbook(file_path)        # 以读写模式打开工作簿
                wb.save(output_file_path)  # 保存工作簿
                wb.close()  # 关闭工作簿
        
        for file_name in os.listdir(Tpp_path):
            if file_name.endswith('.xlsx'):  # 如果是xlsx文件
                file_path = os.path.join(Tpp_path, file_name)  # 构造文件路径
                output_file_path = os.path.join( data_Backup_path, file_name)  # 构造输出文件路径
                wb = load_workbook(file_path)        # 以读写模式打开工作簿
                wb.save(output_file_path)  # 保存工作簿
                wb.close()  # 关闭工作簿


    #删除不包含CJK（中日韩）字元的键值对
    remove_non_cjk(source)
    remove_non_cjk(source_mid)


    keyList=list(source_mid.keys())         #通过字典的keys方法，获取所有的key，转换为list变量
    keyList_len = len(keyList)              #获取原文件key列表的长度，当作于原文的总行数
    print("[INFO] 你的原文长度是",keyList_len)

        #将字典source_mid中的键设为从0开始的整数型数字序号 
    for i in range(keyList_len):        #循环遍历key列表
        source_mid[i] = source_mid.pop(keyList[i])    #将原来的key对应的value值赋给新的key，同时删除原来的key    
    #print("[DEBUG] 你的已修改原文是",source_mid)
  
    result_dict = source_mid.copy() # 先存储未翻译的译文，千万注意不要写等号，不然两个变量会指向同一个内存地址，导致修改一个变量，另一个变量也会被修改
    Translation_Status_List =  [0] * keyList_len   #创建文本翻译状态列表，用于并发时获取每个文本的翻译状态



    #写入过滤和修改key的原文文件，方便debug
    with open(os.path.join(DEBUG_folder, "ManualTransFile_debug.json"), "w", encoding="utf-8") as f:
        json.dump(source_mid, f, ensure_ascii=False, indent=4)

    # ——————————————————————————————————————————构建并发任务池子—————————————————————————————————————————

    # 计算并发任务数
    if keyList_len % Translation_lines == 0:
        tasks_Num = keyList_len // Translation_lines 
    else:
        tasks_Num = keyList_len // Translation_lines + 1


    print("[INFO] 你的翻译任务总数是：", tasks_Num)
    print("\033[1;32m[INFO] \033[0m下面开始进行翻译，请注意保持网络通畅，余额充足", '\n')


    # 创建线程池
    with concurrent.futures.ThreadPoolExecutor (The_Max_workers) as executor:
        # 向线程池提交任务
        for i in range(tasks_Num):
            executor.submit(Make_request)
    # 等待线程池任务完成
        executor.shutdown(wait=True)


    #检查主窗口是否已经退出
    if Running_status == 10 :
        return
    

# ——————————————————————————————————————————检查没能成功翻译的文本，递减行数翻译————————————————————————————————————————

    #计算未翻译文本的数量
    count_not_Translate = Translation_Status_List.count(2)

    #迭代翻译次数
    Number_of_iterations = 0
    #构建递减翻译行数迭代列表   
    Translation_lines_list = divide_by_2345(Translation_lines)

    while count_not_Translate != 0 :
        print("\033[1;33mWarning:\033[0m 仍然有部分未翻译，将进行迭代翻译-----------------------------------")
        print("[INFO] 当前迭代次数：",(Number_of_iterations + 1))
        #将列表变量里未翻译的文本状态初始化
        for i in range(count_not_Translate):      
            if 2 in Translation_Status_List:
                idx = Translation_Status_List.index(2)
                Translation_Status_List[idx] = 0


        
        #根据迭代列表减少翻译行数，直至翻译行数降至1行
        if Number_of_iterations < len(Translation_lines_list):
            Translation_lines = Translation_lines_list[Number_of_iterations]
            # 找到了值，进行后续操作
            print("[INFO] 当前翻译行数设置是：",Translation_lines)
        else:
            # 找不到值，pass
            pass



        # 计算可并发任务总数
        if count_not_Translate % Translation_lines == 0:
            new_count = count_not_Translate // Translation_lines
        else:
            new_count = count_not_Translate // Translation_lines + 1


        # 创建线程池
        with concurrent.futures.ThreadPoolExecutor (The_Max_workers) as executor:
            # 向线程池提交任务
            for i in range(new_count):
                executor.submit(Make_request)
        # 等待线程池任务完成
            executor.shutdown(wait=True)


        #检查主窗口是否已经退出
        if Running_status == 10 :
            return
        
        #检查是否已经陷入死循环
        if Number_of_iterations == 30 :
            break

        #重新计算未翻译文本的数量
        count_not_Translate = Translation_Status_List.count(2) 
        #增加迭代次数，进一步减少翻译行数
        Number_of_iterations = Number_of_iterations + 1

        #如果实时调教功能没有开的话，则每次迭代翻译，增加OpenAI温度,增加随机性
        if Window.Interface18.checkBox.isChecked() == False :
            if OpenAI_temperature + 0.2 <= 1.0 :
                OpenAI_temperature = OpenAI_temperature + 0.2
            else:
                OpenAI_temperature = 1.0
            print("\033[1;33mWarning:\033[0m 当前OpenAI温度是：",OpenAI_temperature)

        #如果只剩下15句左右没有翻译则直接逐行翻译
        if count_not_Translate <= 15:
            Number_of_iterations = len(Translation_lines_list) - 1
            print("\033[1;33mWarning:\033[0m 仅剩下15句未翻译，将进行逐行翻译-----------------------------------")


  # ——————————————————————————————————————————将各类数据处理并保存为各种文件—————————————————————————————————————————

    #处理翻译结果----------------------------------------------------
    new_result_dict = {}
    for i, key in enumerate(source.keys()):     # 使用enumerate()遍历source字典的键，并将其替换到result_dict中
        new_result_dict[key] = result_dict[i]   #在新字典中创建新key的同时把result_dict[i]的值赋予到key对应的值上


    # 将字典存储的译文存储到TrsData.json文件------------------------------------
    if Running_status == 2 :
        #写入文件
        with open(os.path.join(dir_path, "TrsData.json"), "w", encoding="utf-8") as f:
            json.dump(new_result_dict, f, ensure_ascii=False, indent=4)

   # 存储Tpp项目------------------------------------
    else:
        #遍历data_path文件夹里每个的xlsx文件，逐行读取每个文件从A2开始数据，以数据为key，如果source字典中存在该key，则获取value，并将value复制到该行第2列。然后保存文件
        for file_name in os.listdir(data_path):
            if file_name.endswith('.xlsx'):  # 如果是xlsx文件
                file_path = os.path.join(data_path, file_name)  # 构造文件路径
                wb = load_workbook(file_path)  # 以读写模式打开工作簿
                ws = wb.active  # 获取活动工作表
                for row in ws.iter_rows(min_row=2, min_col=1):  # 从第2行开始遍历每一行
                    if len(row) < 2:  # 如果该行的单元格数小于2
                        # 在该行第2列创建一个空单元格
                        new_cell = ws.cell(row=row[0].row, column=2, value="")
                        row = (row[0], new_cell)
                    key = row[0].value  # 获取该行第1列的值作为key
                    #如果key不是None
                    if key is not None:
                        if key in new_result_dict:  # 如果key在new_result_dict字典中
                            value = new_result_dict[key]  # 获取new_result_dict字典中对应的value
                            row[1].value = value  # 将value写入该行第2列
                        else:#如果不在字典中，且第二列没有内容，则复制到第二列中
                            if row[1].value == None:
                                row[1].value = key
                wb.save(file_path)  # 保存工作簿
                wb.close()  # 关闭工作簿



    # —————————————————————————————————————#全部翻译完成——————————————————————————————————————————
    #写入配置保存文件
    Read_Write_Config("write") 

    Ui_signal.update_signal.emit("Translation_completed")#发送信号，激活槽函数,要有参数，否则报错
    print("\n--------------------------------------------------------------------------------------")
    print("\n\033[1;32mSuccess:\033[0m 已完成全部翻译任务，程序已经停止")   
    print("\n\033[1;32mSuccess:\033[0m 请检查译文文件，格式是否错误，存在错行，或者有空行等问题")
    print("\n-------------------------------------------------------------------------------------\n")


# ——————————————————————————————————————————翻译任务线程并发函数(程序核心2)——————————————————————————————————————————
def Make_request():

    global result_dict # 声明全局变量
    global Translation_Status_List  
    global money_used,Translation_Progress,key_list_index,Number_of_requested,Number_of_mark
    global OpenAI_temperature,OpenAI_top_p,OpenAI_frequency_penalty,OpenAI_presence_penalty

    Wrong_answer_count = 0 #错误回答计数，用于错误回答到达一定次数后，取消该任务。

    start_time = time.time()
    timeout = 1200  # 设置超时时间为x秒

    try:#方便排查子线程bug

        # ——————————————————————————————————————————确定翻译位置及段落——————————————————————————————————————————

        #遍历翻译状态列表，找到还没翻译的值和对应的索引位置
        lock1.acquire()  # 获取锁
        for i, status in enumerate(Translation_Status_List):
            if status  == 0:
                start = i     #确定切割开始位置

                if (start + Translation_lines >= keyList_len) :  #确定切割结束位置，注意最后位置是不固定的
                    end = keyList_len  
                else :
                    end = start + Translation_lines
                break
        #修改翻译状态列表位置状态为翻译中
        Translation_Status_List[start:end] = [2] * (end - start)     
        lock1.release()  # 释放锁
        #print("[DEBUG] 当前翻译起始位置是：",start,"------当前翻译结束位置是：", end ) 


        # ——————————————————————————————————————————截取特定段落的文本并进行处理——————————————————————————————————————————

        #读取source_mid源文件中特定起始位置到结束位置的数据,构建新字典变量
        subset_mid = {k: source_mid[k] for k in range( start , end)}     #`k: source_mid[k]`是一个字典键值对，其中`k`表示键，`source_mid[k]`表示该键对应的值。`for k in keys`是一个for循环，它遍历了`keys`列表里的内容，并将其用作字典键。
        #print("[DEBUG] 提取的subset_mid是",subset_mid,'\n','\n') 

        
        #copy前面的代码，将截取文本的键改为从0开始的数字序号，因为AI在回答一万以上的序号时，容易出错
        subset_list=list(subset_mid.keys())        
        subset_len = len(subset_list)              
        for i in range(subset_len):        
            subset_mid[i] = subset_mid.pop(subset_list[i])     
        #print("[DEBUG] 提取的subset_mid是",subset_mid,'\n','\n') 

        #将字典对象编码成 JSON 格式的字符串，方便发送
        subset_str = json.dumps(subset_mid, ensure_ascii=False)    
        #print("[DEBUG] 提取的subset_str是",subset_str,'\n','\n') 

        # ——————————————————————————————————————————整合发送内容——————————————————————————————————————————
        #将JSON 格式的字符串再处理，方便发送            
        d = {"role":"user","content":subset_str}                #将文本整合进字典，符合会话请求格式
        messages = [{"role": "system","content":Prompt}]
        messages.append(d)

        tokens_consume = num_tokens_from_messages(messages, OpenAI_model)  #计算该信息在openai那里的tokens花费

        # ——————————————————————————————————————————开始循环请求，直至成功或失败——————————————————————————————————————————
        while 1 :
            #检查主窗口是否已经退出---------------------------------
            if Running_status == 10 :
                return
            #检查该条消息总tokens数是否大于单条消息最大数量---------------------------------
            if tokens_consume >= (tokens_limit_per-500) :
                print("\033[1;31mError:\033[0m 该条消息总tokens数大于单条消息最大数量" )
                print("\033[1;31mError:\033[0m 该条消息取消任务，进行迭代翻译" )
                break

            #检查子线程是否超时---------------------------------
            if time.time() - start_time > timeout:
                # 超时退出
                print("\033[1;31mError:\033[0m 子线程执行任务已经超时，将暂时取消本次任务")
                break

            #检查请求数量是否达到限制，如果是多key的话---------------------------------
            if len(API_key_list) > 1: #如果存有多个key
                if (Number_of_requested - Number_of_mark) >= 30 :#如果该key请求数已经达到限制次数

                    lock4.acquire()  # 获取锁
                    Number_of_mark = Number_of_requested
                    if (key_list_index + 1) < len(API_key_list):#假如索引值不超过列表最后一个
                            key_list_index = key_list_index + 1 #更换APIKEY索引
                    else :
                            key_list_index = 0

                    #更新API
                    #openai.api_key = API_key_list[key_list_index]
                    on_update_signal("CG_key")

                    #重置频率限制，重置请求时间
                    api_tokens.tokens = tokens_limit_per * 2
                    api_request.last_request_time = 0

                    print("\033[1;33mWarning:\033[0m 该key请求数已达30,将进行KEY的更换")
                    print("\033[1;33mWarning:\033[0m 将API-KEY更换为第",key_list_index+1,"个 , 值为：", API_key_list[key_list_index] ,'\n')
                    lock4.release()  # 释放锁

            # 检查子是否符合速率限制---------------------------------
            if api_tokens.consume(tokens_consume * 2 ) and api_request.send_request():

                #如果能够发送请求，则扣除令牌桶里的令牌数
                api_tokens.tokens = api_tokens.tokens - (tokens_consume * 2 )

                print("[INFO] 已发送请求,正在等待AI回复中--------------")
                print("[INFO] 已进行请求的次数：",Number_of_requested)
                print("[INFO] 花费tokens数预计值是：",tokens_consume * 2) 
                print("[INFO] 桶中剩余tokens数是：", api_tokens.tokens // 1)
                print("[INFO] 当前发送内容：\n", messages ,'\n','\n')

                # ——————————————————————————————————————————开始发送会话请求——————————————————————————————————————————
                try:
                    lock5.acquire()  # 获取锁
                    Number_of_requested = Number_of_requested + 1#记录请求数
                    #如果启用实时参数设置
                    if Window.Interface18.checkBox.isChecked() :
                        #获取界面配置信息
                        OpenAI_temperature = Window.Interface18.slider1.value() * 0.1
                        OpenAI_top_p = Window.Interface18.slider2.value() * 0.1
                        OpenAI_frequency_penalty = Window.Interface18.slider3.value() * 0.1
                        OpenAI_presence_penalty = Window.Interface18.slider4.value() * 0.1
                        #输出到控制台
                        print("[INFO] 实时参数设置已启用")
                        print("[INFO] 当前temperature是:",OpenAI_temperature)
                        print("[INFO] 当前top_p是:",OpenAI_top_p)
                        print("[INFO] 当前frequency_penalty是:",OpenAI_frequency_penalty)
                        print("[INFO] 当前presence_penalty是:",OpenAI_presence_penalty,'\n','\n')
                    lock5.release()  # 释放锁
                    response = openai.ChatCompletion.create(
                        model= OpenAI_model,
                        messages = messages ,
                        temperature=OpenAI_temperature,
                        top_p = OpenAI_top_p,
                        frequency_penalty=OpenAI_frequency_penalty,
                        presence_penalty=OpenAI_presence_penalty
                        )

                #一旦有错误就抛出错误信息，一定程度上避免网络代理波动带来的超时问题
                except Exception as e:
                    print("\033[1;33m线程ID:\033[0m ", threading.get_ident())
                    print("\033[1;31mError:\033[0m api请求出现问题！错误信息如下")
                    print(f"Error: {e}\n")
                    #处理完毕，再次进行请求
                    continue


                #——————————————————————————————————————————收到回复，并截取回复内容中的文本内容 ————————————————————————————————————————       
                response_content = response['choices'][0]['message']['content'] 


                #截取回复内容中返回的tonkens花费，并计算金钱花费
                lock3.acquire()  # 获取锁

                prompt_tokens_used = int(response["usage"]["prompt_tokens"]) #本次请求花费的tokens
                completion_tokens_used = int(response["usage"]["completion_tokens"]) #本次回复花费的tokens
                total_tokens_used = int(response["usage"]["total_tokens"]) #本次请求+回复花费的tokens


                Request_Costs  = prompt_tokens_used * Request_Pricing  #本次请求花费的金钱
                Response_Costs = completion_tokens_used * Response_Pricing #本次回复花费的金钱
                The_round_trip_cost = Request_Costs + Response_Costs #本次往返花费的金钱


                money_used = money_used + The_round_trip_cost #累计花费的金钱

                lock3.release()  # 释放锁

                print("[INFO] 已成功接受到AI的回复--------------")
                print("[INFO] 此次请求消耗的总tokens：",total_tokens_used )
                print("[INFO] 此次请求往返的总金额：",The_round_trip_cost )
                print("[INFO] AI回复的文本内容：\n",response_content ,'\n','\n')

             # ——————————————————————————————————————————对AI回复内容进行各种处理和检查——————————————————————————————————————————


                #专门针对 (There is no need to translate this text as it does not contain any Japanese characters.) 这种情况进行处理
                if response_content[-1]  ==  ')':                   # 再检查 response_check 的最后一个字符是不是括号
                    pos = response_content.rfind('(')                  # 从后往前查找最后一个前括号的位置
                    response_content = response_content[:pos]           # 删除前括号号及其后面的所有字符


                Error_Type = [0,0,0,0]   #错误类型存储列表
                print("[INFO] 开始对AI回复内容进行各项检查--------------") 

                #检查回复内容的json格式------------------------------------------------------ 
                try:
                    response_content_dict = json.loads(response_content) #注意转化为字典的数字序号key是字符串类型           
                except :                                            
                    Error_Type[0] = 1

                #主要检查AI回复时，键值对数量对不对------------------------------------------------------

                if Error_Type[0] == 0:
                    if(len(response_content_dict)  !=  (end - start ) ):    
                        Error_Type[1] = 1


                #主要检查AI回复时，有没有某一行为空或者只是回复符号------------------------------------------------------
                if (Error_Type[0]== 0) and (Error_Type[1] == 0): #注意错误的缩写方法Error_Type[0] or Error_Type[1] == 0，以及注意大括号括起来下的整体逻辑
                    for value in response_content_dict.values():
                        #检查value是不是None，因为AI回回复null，但是json.loads()会把null转化为None
                        if value is None:
                            Error_Type[2] = 1
                            break

                        # 检查value是不是空字符串，因为AI回回复空字符串，但是json.loads()会把空字符串转化为""
                        if value == "":
                            Error_Type[2] = 1
                            break
                        #统计回复内容中的中文、日文、韩文、字符数量
                        A,B,C,D= count_japanese_chinese_korean(value)

                        #如果有某一行只是回复符号就把Error_Type[2]改为1
                        if A+B+C+D == 0:
                            Error_Type[2] = 1
                            break

                #主要检查AI回复时，符号与字数是否能够与原文大致对上------------------------------------------------------
                if (Error_Type[0]== 0) and (Error_Type[1]== 0) and (Error_Type[2] == 0):
                    Check_dict = {}
                    for i in range(len(subset_mid)):
                        Check_dict[subset_mid[i]] = response_content_dict[str(i)]


                    #计算Check_dict中的键值对的个数，并创建列表来存储键值对的错误状态
                    pairs_count = len(Check_dict)
                    error_list = [1] * pairs_count

                    i = 0#循环计次，顺便改变错误状态列表索引位置

                    for k, v in Check_dict.items():
                        error_count = 0
                            
                        # 用正则表达式匹配原文与译文中的标点符号
                        k_syms = re.findall(r'[。！？…♡♥=★]', k)
                        v_syms = re.findall(r'[。！？…♡♥=★]', v)

                        #假如v_syms与k_syms都不为空
                        if len(v_syms) != 0 and len(k_syms) != 0:
                            #计算v_syms中的元素在k_syms中存在相同元素的比例
                            P = len([sym for sym in v_syms if sym in k_syms]) / len(v_syms)
                        #假如v_syms与k_syms都为空，即原文和译文都没有标点符号
                        elif len(v_syms) == 0 and len(k_syms) == 0:
                            P = 1
                        else:
                            P = 0
                        #如果标点符号的比例相差较大，则错误+1
                        if P < 0.5:
                            error_count += 1



                        #计算k中的日文、中文,韩文，英文字母的个数
                        Q,W,E,R = count_japanese_chinese_korean(k)
                        #计算v中的日文、中文,韩文，英文字母的个数
                        A,S,D,F = count_japanese_chinese_korean(v)
                        #如果日文、中文的个数相差较大，则错误+1
                        if abs((Q+W+E+R) - (A+S+D+F)) > 8: 
                            error_count += 1



                        #如果error_count为2
                        if error_count == 2:
                            #当前位置的状态在状态列表中改为0，并改变error_list中的值和相邻元素的值为0
                            error_list[i] = 0
                            if i != 0:
                                error_list[i-1] = 0
                            if i != pairs_count - 1:
                                error_list[i+1] = 0

                        #该次循环结束，位置索引+1
                        i = i + 1

                    #遍历完成，统计error_list列表中值为0的个数占总个数的比例，并转化为百分数
                    error_list_count = error_list.count(0)
                    error_list_count_percent = error_list_count / pairs_count * 100
                    error_list_count_percent = round(error_list_count_percent, 2)

                    #如果错误的比例大于阈值，则错误
                    Error_Threshold = 40
                    if error_list_count_percent >= Error_Threshold:
                        Error_Type[3] = 1

                    #如果翻译行数已经迭代到了10行，就忽略错误，避免死循环
                    if end - start == 10:
                        Error_Type[3] = 0


                #如果出现回复错误------------------------------------------------------
                if (Error_Type[0]== 1)  or (Error_Type[1]== 1) or (Error_Type[2]== 1) or (Error_Type[3]  == 1) :
                    if Error_Type[0] == 1 :
                        print("\033[1;33mWarning:\033[0m AI回复内容不符合json格式,将进行重新翻译\n")
                        Error_message = "Warning: AI回复内容不符合json格式要求,将进行重新翻译\n"
                    elif Error_Type[1] == 1 :
                        print("\033[1;33mWarning:\033[0m AI回复内容键值对数量与原来数量不符合,将进行重新翻译\n")
                        Error_message = "Warning: AI回复内容键值对数量与原来数量不符合,将进行重新翻译\n"
                    elif Error_Type[2] == 1 :
                        print("\033[1;33mWarning:\033[0m AI回复内容中有空行或仅符号,将进行重新翻译\n")
                        Error_message = "Warning: AI回复内容中有空行或仅符号,将进行重新翻译\n"
                    elif Error_Type[3] == 1 :
                        print("\033[1;33mWarning:\033[0m AI回复内容的符号与字数与原文的不符合程度为:",error_list_count_percent,"%,大于",Error_Threshold,"%阈值，将进行重新翻译\n")
                        Error_message = "Warning: AI回复内容的符号与字数与原文不符合大于阈值,将进行重新翻译\n"

                    #错误回复计次
                    Wrong_answer_count = Wrong_answer_count + 1
                    print("\033[1;33mWarning:\033[0m AI回复内容格式错误次数:",Wrong_answer_count,"到达3次后将该段文本进行迭代翻译\n")

                    #将错误回复和原文文本写入DEBUG文件夹，以便修复BUG
                    if  Wrong_answer_count == 1 :#当第一次出现错误回复时
                        # 创建专属文件夹路径
                        The_folder_name = "Wrong position  "+str(start) + "——" +str(end)
                        folder_path = os.path.join(DEBUG_folder, The_folder_name)
                        os.makedirs(folder_path, exist_ok=True)

                        #写入原文文本，方便做对比
                        with open( os.path.join(folder_path, "Original text.json"), "w", encoding="utf-8") as f:
                            json.dump(subset_mid, f, ensure_ascii=False, indent=4)

                        #创建存储错误回复的变量
                        Error_text_str = ""
                    
                    if Wrong_answer_count >= 1 :#当从第一次出现错误回复开始，每次都
                        #收集错误的回复内容，并写入文件
                        Error_text_str = Error_text_str +'\n' + response_content +'\n' + Error_message +'\n'
                        with open( os.path.join(folder_path, "Error text.txt"), "w", encoding="utf-8") as f:
                            f.write(Error_text_str)

                    #检查回答错误次数，如果达到限制，则跳过该句翻译。
                    if Wrong_answer_count >= 3 :
                        print("\033[1;33mWarning:\033[0m 错误次数已经达限制,将该段文本进行迭代翻译！\n")    
                        break


                    #进行下一次循环
                    time.sleep(1)                 
                    continue

                #如果没有出现错误------------------------------------------------------ 
                else:
                    
                    print("[INFO] AI回复内容字符串符合JSON 格式")
                    print("[INFO] AI回复内容键值对数量符合要求")
                    print("[INFO] AI回复内容中没有空行或仅符号")
                    print("[INFO] AI回复内容的符号与字数与原文的不符合程度为:",error_list_count_percent,"%,小于",Error_Threshold,"%阈值\n")

                    #格式检查通过，将AI酱回复的内容数字序号进行修改，方便后面进行读写json文件
                    new_response = re.sub(r'"(\d+)"', lambda x: '"' + str(int(x.group(1))+start) + '"', response_content)


                    lock1.acquire()  # 获取锁
                    #修改文本翻译状态列表的状态，把这段文本修改为已翻译
                    Translation_Status_List[start:end] = [1] * (end - start) 

                    Translation_Progress = Translation_Status_List.count(1) / keyList_len  * 100
                    Ui_signal.update_signal.emit("Update_ui")#发送信号，激活槽函数,要有参数，否则报错
                    lock1.release()  # 释放锁
                    print(f"\n--------------------------------------------------------------------------------------")
                    print(f"\n\033[1;32mSuccess:\033[0m 翻译已完成：{Translation_Progress:.2f}%               已花费费用：{money_used:.4f}＄")
                    print(f"\n--------------------------------------------------------------------------------------\n")



                    lock2.acquire()  # 获取锁
                    # 用字典类型存储每次请求的译文
                    new_response_dict =json.loads(new_response )
                    for key, value in new_response_dict.items():# 遍历new_response_dict中的键值对
                        # 判断key是否在result_dict中出现过，注意两个字典的key变量类型是不同的
                        if int(key) in result_dict:
                            # 如果出现过，则将result_dict中对应键的值替换为new_response_dict中对应键的值
                            result_dict[int(key)] = value
 
                    #备份翻译数据
                    File_Backup()

                    lock2.release()  # 释放锁



                    break



   #子线程抛出错误信息
    except Exception as e:
        print("\033[1;31mError:\033[0m 线程出现问题！错误信息如下")
        print(f"Error: {e}\n")
        return


# ——————————————————————————————————————————检查词义错误单独功能函数——————————————————————————————————————————
def Check_wrong():
    global file_name,dir_path,Backup_folder ,Translation_lines,Running_status,The_Max_workers,DEBUG_folder
    global keyList_len ,   Translation_Status_List , money_used,source,source_mid,result_dict,Translation_Progress
    # ——————————————————————————————————————————清空进度,花销与初始化变量存储的内容—————————————————————————————————————————

    money_used = 0
    Translation_Progress = 0 

    result_dict = {}
    source = {}  # 存储字符串数据的字典

    # 创建DEBUG文件夹路径
    DEBUG_folder = os.path.join(dir_path, 'DEBUG Folder')
    #使用`os.makedirs()`函数创建新文件夹，设置`exist_ok=True`参数表示如果文件夹已经存在，不会抛出异常
    os.makedirs(DEBUG_folder, exist_ok=True)

    # 创建备份文件夹路径
    Backup_folder = os.path.join(dir_path, 'Backup Folder')
    #使用`os.makedirs()`函数创建新文件夹，设置`exist_ok=True`参数表示如果文件夹已经存在，不会抛出异常
    os.makedirs(Backup_folder, exist_ok=True) 
    # ——————————————————————————————————————————读取原文文件并处理—————————————————————————————————————————

    if Running_status == 4:
        with open(file_name, 'r',encoding="utf-8") as f:               
            source_str = f.read()       #读取原文文件，以字符串的形式存储，直接以load读取会报错

            result_dict = json.loads(source_str) #转换为字典类型的变量source，当作最后翻译文件的原文源


    elif Running_status == 5:
        # 遍历文件夹中的所有xlsx文件到source变量里
        for file_name in os.listdir(Tpp_path):
            if file_name.endswith('.xlsx'):  # 如果是xlsx文件
                file_path = os.path.join(Tpp_path, file_name)  # 构造文件路径
                wb = load_workbook(file_path, read_only=True)  # 以只读模式打开工作簿
                ws = wb.active  # 获取活动工作表
                for row in ws.iter_rows(min_row=2, min_col=1):  # 从第2行开始遍历每一行
                    #如果第1列的值不为空，过滤掉空行
                    if row[0].value is not None:
                        key = row[0].value  # 获取该行第1列的值作为key
                        value = row[1].value  # 获取该行第2列的值作为value
                        result_dict[key] = value  # 将key和value添加到字典source中
                wb.close()  # 关闭工作簿

        #在输出文件夹里新建文件夹data
        data_path = os.path.join(dir_path, 'data')
        os.makedirs(data_path, exist_ok=True)

        #在备份文件夹里新建文件夹data
        data_Backup_path = os.path.join(Backup_folder, 'data')
        os.makedirs(data_Backup_path, exist_ok=True)

        #复制原项目data文件夹所有文件到输出文件夹data文件夹里和备份文件夹的data里面
        for file_name in os.listdir(Tpp_path):
            if file_name.endswith('.xlsx'):  # 如果是xlsx文件
                file_path = os.path.join(Tpp_path, file_name)  # 构造文件路径
                output_file_path = os.path.join(data_path, file_name)  # 构造输出文件路径
                wb = load_workbook(file_path)        # 以读写模式打开工作簿
                wb.save(output_file_path)  # 保存工作簿
                wb.close()  # 关闭工作簿
        
        for file_name in os.listdir(Tpp_path):
            if file_name.endswith('.xlsx'):  # 如果是xlsx文件
                file_path = os.path.join(Tpp_path, file_name)  # 构造文件路径
                output_file_path = os.path.join( data_Backup_path, file_name)  # 构造输出文件路径
                wb = load_workbook(file_path)        # 以读写模式打开工作簿
                wb.save(output_file_path)  # 保存工作簿
                wb.close()  # 关闭工作簿



    #将result_dict的key作为source的key，并复制source的key的值为该key对应的value
    source = result_dict.copy()
    #将source的value的值全部替换为key的值，这样source的key和value就一样了
    for key, value in source.items():
        source[key] = key

    source_mid = source.copy()  # 复制source的值到source_mid，作为中间变量



    keyList=list(source_mid.keys())         #通过字典的keys方法，获取所有的key，转换为list变量
    keyList_len = len(keyList)              #获取原文件key列表的长度，当作于原文的总行数
    print("[INFO] 你的原文长度是",keyList_len)

    #将字典source_mid中的键设为从0开始的整数型数字序号 
    for i in range(keyList_len):        #循环遍历key列表
        source_mid[i] = source_mid.pop(keyList[i])    #将原来的key对应的value值赋给新的key，同时删除原来的key    
    #print("[DEBUG] 你的已修改原文是",source_mid)


    #将字典result_dict中的键设为从0开始的整数型数字序号 
    for i in range(keyList_len):        #循环遍历key列表
        result_dict[i] = result_dict.pop(keyList[i])    #将原来的key对应的value值赋给新的key，同时删除原来的key    
    #print("[DEBUG] 你的已修改原文是",result_dict)
  

    Translation_Status_List =  [1] * keyList_len   #创建文本翻译状态列表，用于并发时获取每个文本的翻译状态


    # ——————————————————————————————————————————进行语义相似度检查，并重翻译—————————————————————————————————————————
    #进行语义相似度检查----------------------------------------------------
    print("\033[1;33mWarning:\033[0m 正在检查译文中翻译错误的内容，请耐心等待-----------------------------------")

    T2T_model = SentenceTransformer('sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2')  #这个模型快点
    sentences = ["", ""]  #这里是两个空字符串，后面会被替换

    #存储错误文本的字典
    error_txt_dict = {}
    #创建存储错误文本的文件夹
    ErrorTxt_folder = os.path.join(DEBUG_folder, 'ErrorTxt Folder')
    #使用`os.makedirs()`函数创建新文件夹，设置`exist_ok=True`参数表示如果文件夹已经存在，不会抛出异常
    os.makedirs(ErrorTxt_folder, exist_ok=True)
        
    #错误文本计数变量
    count_error = 0

    #循环检测文本，如果语义相似度小于阈值，则将 Translation_Status_List[i]中的数值改为0，表示需要重翻
    for i, key in enumerate(result_dict.keys()):
        sentences[0] = source_mid[key]
        sentences[1] = result_dict[key]

        #检测sentence[0]与sentence[1]是不是为null，如果是null，则跳过，因为null是无法计算语义相似度的，而且报错，主要因为AI回复时会出现null回答
        if sentences[0] == "" or sentences[1] == "":
            Translation_Status_List[i]  = 0
            count_error = count_error + 1
            print("[INFO] 因为AI回复时没有内容，出现了为NUll型数据，需要重翻译")
            print("\033[1;33mWarning:\033[0m 当前错误文本数量：", count_error)
            continue

        #将sentence[0]与sentence[1]转换成字符串数据，确保能够被语义相似度检查模型识别，防止数字型数据导致报错
        sentences[0] = str(sentences[0])
        sentences[1] = str(sentences[1])

        #计算语义相似度
        cosine_scores = util.pytorch_cos_sim(T2T_model.encode(sentences[0]), T2T_model.encode(sentences[1]))
        #cos_sim = vec1.dot(vec2) / (np.linalg.norm(vec1) * np.linalg.norm(vec2))

        #输出sentence里的两个文本 和 语义相似度检查结果
        print("[INFO] 原文是：", sentences[0])
        print("[INFO] 译文是：", sentences[1])

        #将语义相似度转换为百分比
        percentage = cosine_scores.item() * 100
        #如果语义相似度小于于等于阈值，则将 Translation_Status_List[i]中的数值改为0，表示需要重翻译
        if percentage <= 50:
            Translation_Status_List[i]  = 0
            count_error = count_error + 1
            print("[INFO] 语义相似度检查结果：", percentage, "%", "，需要重翻译")
            #错误文本计数提醒
            print("\033[1;33mWarning:\033[0m 当前错误文本数量：", count_error)

            #将错误文本存储到字典里
            error_txt_dict[sentences[0]] = sentences[1]


        else :
            print("[INFO] 语义相似度检查结果：", percentage, "%", "，不需要重翻译")
            
        #输出遍历进度，转换成百分百进度
        print("[INFO] 当前检查进度：", round((i+1)/len(result_dict.keys())*100,2), "%")

    #将错误文本字典写入json文件
    with open(os.path.join(ErrorTxt_folder, "error_txt_dict.json"), 'w', encoding='utf-8') as f:
        json.dump(error_txt_dict, f, ensure_ascii=False, indent=4)

                
            
    #重新翻译需要重翻译的文本----------------------------------------------------
    print("\033[1;33mWarning:\033[0m 针对错误译文进行重新翻译-----------------------------------")

    #计算需要翻译文本的数量
    count_not_Translate = Translation_Status_List.count(0)
    #设置为逐行翻译
    Translation_lines = 1

    #记录翻译次数
    Number_of_iterations = 0

    while count_not_Translate != 0 :
        #将列表变量里未翻译的文本状态初始化
        for i in range(count_not_Translate):      
            if 2 in Translation_Status_List:
                idx = Translation_Status_List.index(2)
                Translation_Status_List[idx] = 0

        # 计算可并发任务总数
        if count_not_Translate % Translation_lines == 0:
            new_count = count_not_Translate // Translation_lines
        else:
            new_count = count_not_Translate // Translation_lines + 1

        # 创建线程池
        with concurrent.futures.ThreadPoolExecutor (The_Max_workers) as executor:
            # 向线程池提交任务
            for i in range(new_count):
                executor.submit(Make_request)
        # 等待线程池任务完成
            executor.shutdown(wait=True)

        #检查主窗口是否已经退出
        if Running_status == 10 :
            return
            
                    
        #检查是否已经陷入死循环
        if Number_of_iterations == 10 :
            print("\033[1;33mWarning:\033[0m 已达到最大循环次数，退出翻译任务，不影响后续使用-----------------------------------")
            break

        #重新计算未翻译文本的数量
        count_not_Translate = Translation_Status_List.count(2)+ Translation_Status_List.count(0)

        #记录循环次数
        Number_of_iterations = Number_of_iterations + 1
        print("\033[1;33mWarning:\033[0m 当前循环翻译次数：", Number_of_iterations, "次    到达最大循环次数10次后将退出翻译任务-------------------------------")

    print("\033[1;33mWarning:\033[0m 已重新翻译完成-----------------------------------")


    # ——————————————————————————————————————————将各类数据处理并保存为各种文件—————————————————————————————————————————

    #处理翻译结果----------------------------------------------------
    new_result_dict = {}
    for i, key in enumerate(source.keys()):     # 使用enumerate()遍历source字典的键，并将其替换到result_dict中
        new_result_dict[key] = result_dict[i]   #在新字典中创建新key的同时把result_dict[i]的值赋予到key对应的值上


    # 将字典存储的译文存储到TrsData.json文件------------------------------------
    if Running_status == 4 :
        #写入文件
        with open(os.path.join(dir_path, "TrsData.json"), "w", encoding="utf-8") as f:
            json.dump(new_result_dict, f, ensure_ascii=False, indent=4)

   # 存储Tpp项目------------------------------------
    elif Running_status == 5 :
        #遍历data_path文件夹里每个的xlsx文件，逐行读取每个文件从A2开始数据，以数据为key，如果source字典中存在该key，则获取value，并将value复制到该行第2列。然后保存文件
        for file_name in os.listdir(data_path):
            if file_name.endswith('.xlsx'):  # 如果是xlsx文件
                file_path = os.path.join(data_path, file_name)  # 构造文件路径
                wb = load_workbook(file_path)  # 以读写模式打开工作簿
                ws = wb.active  # 获取活动工作表
                for row in ws.iter_rows(min_row=2, min_col=1):  # 从第2行开始遍历每一行
                    if len(row) < 2:  # 如果该行的单元格数小于2
                        new_cell = ws.cell(row=row[0].row, column=2, value="")
                        row = (row[0], new_cell)

                    key = row[0].value  # 获取该行第1列的值作为key
                    #如果key不是None
                    if key is not None:
                        if key in source:  # 如果key在source字典中
                            value = new_result_dict[key]  # 获取source字典中对应的value
                            row[1].value = value  # 将value写入该行第2列
                        else:#如果不在字典中，则复制到第二列中
                            row[1].value = key
                wb.save(file_path)  # 保存工作簿
                wb.close()  # 关闭工作簿



    # —————————————————————————————————————#全部翻译完成——————————————————————————————————————————
    #写入配置保存文件
    Read_Write_Config("write") 

    Ui_signal.update_signal.emit("Translation_completed")#发送信号，激活槽函数,要有参数，否则报错
    print("\n--------------------------------------------------------------------------------------")
    print("\n\033[1;32mSuccess:\033[0m 已完成全部翻译任务，程序已经停止")   
    print("\n\033[1;32mSuccess:\033[0m 请检查译文文件，格式是否错误，存在错行，或者有空行等问题")
    print("\n-------------------------------------------------------------------------------------\n")


# ——————————————————————————————————————————下面都是UI相关代码——————————————————————————————————————————

class Widget11(QFrame):#自定义的widget内容界面

    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用



        #设置各个控件-----------------------------------------------------------------------------------------

        #设置基础参数
        x = 60
        y = 400


        #设置“启用该账号”标签
        self.label5 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label5.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label5.setText("启用该平台")
        self.label5.move(x, y-100)

        #设置“启用该账号”开关
        self.checkBox = CheckBox('OpenAI官方', self)
        self.checkBox.move(x, y-60)
        self.checkBox.stateChanged.connect(self.checkBoxChanged)



        #设置“账号类型”标签
        self.label2 = QLabel(parent = self, flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")#设置字体，大小，颜色
        self.label2.setText("账号类型")
        self.label2.move(x, y)

        #设置“账号类型”下拉选择框
        self.comboBox = ComboBox(self) #以demo为父类
        self.comboBox.addItems(['免费账号', '付费账号(48h内)', '付费账号(48h后)'])
        self.comboBox.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox.setFixedSize(150, 30)
        self.comboBox.move(x, y+40)


        #设置“模型选择”标签
        self.label3 = QLabel(parent = self, flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")#设置字体，大小，颜色
        self.label3.setText("模型选择")
        self.label3.move(x+400, y)

        #设置“模型类型”下拉选择框
        self.comboBox2 = ComboBox(self) #以demo为父类
        self.comboBox2.addItems(['gpt-3.5-turbo', 'gpt-4'])
        self.comboBox2.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox2.setFixedSize(150, 30)
        self.comboBox2.move(x+400, y+40)


        #设置“代理地址”标签
        self.label4 = QLabel(parent = self, flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")#设置字体，大小，颜色
        self.label4.setText("代理地址")
        self.label4.move(x, y+100)

        #设置“代理地址”的输入框
        self.LineEdit1 = LineEdit(self)
        self.LineEdit1.move(x, y+140)
        self.LineEdit1.setFixedSize(700, 30)
        #self.LineEdit1.setText("http://127.0.0.1:10080")

        #设置“API KEY”标签
        self.label5 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label5.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label5.setText("API KEY")
        self.label5.move(x, y+200)

        #设置“API KEY”的输入框
        self.TextEdit2 = TextEdit(self)
        self.TextEdit2.move(x, y+240)
        self.TextEdit2.setFixedSize(700, 30)
        #self.TextEdit2.setInputMethodHints(Qt.ImhNoAutoUppercase)




        #设置“测试请求”的按钮
        self.primaryButton1 = PrimaryPushButton('测试请求', self, FIF.SEND)
        self.primaryButton1.move(x+280, y+300)
        self.primaryButton1.clicked.connect(On_button_clicked4) #按钮绑定槽函数


    def checkBoxChanged(self, isChecked: bool):
        global Running_status
        if isChecked :
            Window.Interface12.checkBox.setChecked(False)
            CreateSuccessInfoBar("已设置使用OpenAI官方进行翻译")


class Widget12(QFrame):#自定义的widget内容界面


    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用



        #设置各个控件-----------------------------------------------------------------------------------------

        #设置基础参数
        x = 60
        y = 400


        #设置“启用该账号”标签
        self.label5 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label5.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label5.setText("启用该平台")
        self.label5.move(x, y-100)

        #设置“启用该账号”开关
        self.checkBox = CheckBox('OpenAI代理', self)
        self.checkBox.move(x, y-60)
        self.checkBox.stateChanged.connect(self.checkBoxChanged)


        #设置“账号类型”标签
        self.label2 = QLabel(parent = self, flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")#设置字体，大小，颜色
        self.label2.setText("账号类型")
        self.label2.move(x, y)

        #设置“账号类型”下拉选择框
        self.comboBox = ComboBox(self) #以demo为父类
        self.comboBox.addItems(['代理账号'])
        self.comboBox.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox.setFixedSize(150, 30)
        self.comboBox.move(x, y+40)


        #设置“模型选择”标签
        self.label3 = QLabel(parent = self, flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")#设置字体，大小，颜色
        self.label3.setText("模型选择")
        self.label3.move(x+400, y)

        #设置“模型类型”下拉选择框
        self.comboBox2 = ComboBox(self) #以demo为父类
        self.comboBox2.addItems(['gpt-3.5-turbo', 'gpt-4'])
        self.comboBox2.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox2.setFixedSize(150, 30)
        self.comboBox2.move(x+400, y+40)


        #设置“API代理地址”标签
        self.label4 = QLabel(parent = self, flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")#设置字体，大小，颜色
        self.label4.setText("域名地址")
        self.label4.move(x, y+100)

        #设置“API代理地址”的输入框
        self.LineEdit1 = LineEdit(self)
        self.LineEdit1.move(x, y+140)
        self.LineEdit1.setFixedSize(700, 30)
        self.LineEdit1.setText("https://api.openai-sb.com/v1")

        #设置“API KEY”标签
        self.label5 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label5.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label5.setText("API KEY")
        self.label5.move(x, y+200)

        #设置“API KEY”的输入框
        self.TextEdit2 = TextEdit(self)
        self.TextEdit2.move(x, y+240)
        self.TextEdit2.setFixedSize(700, 30)
        #self.TextEdit2.setInputMethodHints(Qt.ImhNoAutoUppercase)




        #设置“测试请求”的按钮
        self.primaryButton1 = PrimaryPushButton('测试请求', self, FIF.SEND)
        self.primaryButton1.move(x+280, y+300)
        self.primaryButton1.clicked.connect(On_button_clicked4) #按钮绑定槽函数


    def checkBoxChanged(self, isChecked: bool):
        global Running_status
        if isChecked :
            Window.Interface11.checkBox.setChecked(False)
            CreateSuccessInfoBar("已设置使用OpenAI国内代理平台进行翻译")


class Widget15(QFrame):#自定义的widget内容界面

    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用


        #设置各个控件-----------------------------------------------------------------------------------------

        #设置基础参数
        x = 60
        y = 100

        #设置“翻译行数”标签
        self.label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label7.setText("Lines")
        self.label7.move(x, y)


       #设置“翻译行数”数值输入框
        self.spinBox1 = SpinBox(self)    
        self.spinBox1.move(x, y+40)
        self.spinBox1.setValue(40)


        #设置“语义检查”标签
        self.label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label7.setText("语义检查")
        self.label7.move(x+400, y)
        self.label7.hide()

       #设置“语义检查”选择开关
        self.SwitchButton1 = SwitchButton(parent=self)    
        self.SwitchButton1.move(x+400, y+40)
        self.SwitchButton1.checkedChanged.connect(self.onCheckedChanged)
        self.SwitchButton1.hide()


        #设置“Prompt”标签
        self.label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label7.setText("Prompt")
        self.label7.move(x, y+100)

        #设置“Prompt”的输入框
        self.TextEdit = TextEdit(self)
        self.TextEdit.move(x, y+140)
        self.TextEdit.setFixedSize(700, 200)
        self.TextEdit.setText(Prompt)


        #设置“文件位置”标签
        self.label8 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label8.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label8.setText("文件位置")
        self.label8.move(x, y+380)

        #设置打开文件按钮
        self.pushButton1 = PushButton('选择文件', self, FIF.DOCUMENT)
        self.pushButton1.move(x, y+420)
        self.pushButton1.clicked.connect(On_button_clicked1) #按钮绑定槽函数

        #设置“文件位置”显示
        self.label9 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label9.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label9.resize(500, 20)#设置标签大小
        self.label9.setText("请选择需要翻译的json文件")
        self.label9.move(x+150, y+425)   


        #设置“输出文件夹”标签
        self.label10 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label10.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label10.setText("输出文件夹")
        self.label10.move(x, y+480)  

        #设置输出文件夹按钮
        self.pushButton2 = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton2.move(x, y+520)
        self.pushButton2.clicked.connect(On_button_clicked3) #按钮绑定槽函数

        #设置“输出文件夹”显示
        self.label11 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label11.resize(500, 20)
        self.label11.setText("请选择翻译文件存储文件夹")
        self.label11.move(x+150, y+525)    



        #设置“开始翻译”的按钮
        self.primaryButton1 = PrimaryPushButton('开始翻译', self, FIF.UPDATE)
        self.primaryButton1.move(x+280, y+600)
        self.primaryButton1.clicked.connect(On_button_clicked5) #按钮绑定槽函数


        #设置“已花费”标签
        self.label12 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label12.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label12.setText("已花费")
        self.label12.move(x, y+640)
        self.label12.hide()  #先隐藏控件

        #设置“已花费金额”具体标签
        self.label13 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label13.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label13.resize(500, 20)#设置标签大小
        self.label13.setText("0＄")
        self.label13.move(x+60, y+640)
        self.label13.hide()  #先隐藏控件

        #设置翻译进度条控件
        self.progressBar = QProgressBar(self)
        self.progressBar.setMinimum(0)
        self.progressBar.setMaximum(100)
        self.progressBar.setValue(0)
        self.progressBar.setFixedSize(700, 30)
        self.progressBar.move(x, y+670)
        self.progressBar.setStyleSheet("QProgressBar::chunk { text-align: center; } QProgressBar { text-align: left; }")#使用setStyleSheet()方法设置了进度条块的文本居中对齐，并且设置了进度条的文本居左对齐
        self.progressBar.setFormat("已翻译: %p%")
        self.progressBar.hide()  #先隐藏控件

    def onCheckedChanged(self, isChecked: bool):
        if isChecked :
            self.SwitchButton1.setText("已开启")
            CreateWarningInfoBar("第一次使用语义检查功能，会下载500mb左右的模型，请注意网络环境与电脑存储空间！")
        else :
            self.SwitchButton1.setText("已关闭")


class Widget16(QFrame):#自定义的widget内容界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用


        #设置各个控件-----------------------------------------------------------------------------------------

        #设置基础参数
        x = 60
        y = 100

        #设置“翻译行数”标签
        self.label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label7.setText("Lines")
        self.label7.move(x, y)


       #设置“翻译行数”数值输入框
        self.spinBox1 = SpinBox(self)    
        self.spinBox1.move(x, y+40)
        self.spinBox1.setValue(40)



        #设置“语义检查”标签
        self.label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label7.setText("语义检查")
        self.label7.move(x+400, y)
        self.label7.hide()  #先隐藏控件


       #设置“语义检查”选择开关
        self.SwitchButton1 = SwitchButton(parent=self)    
        self.SwitchButton1.move(x+400, y+40)
        self.SwitchButton1.checkedChanged.connect(self.onCheckedChanged)
        self.SwitchButton1.hide()  #先隐藏控件


        #设置“Prompt”标签
        self.label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label7.setText("Prompt")
        self.label7.move(x, y+100)

        #设置“Prompt”的输入框
        self.TextEdit = TextEdit(self)
        self.TextEdit.move(x, y+140)
        self.TextEdit.setFixedSize(700, 200)
        self.TextEdit.setText(Prompt)


        #设置“项目文件夹”标签
        self.label8 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label8.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label8.setText("项目文件夹")
        self.label8.move(x, y+380)

        #设置打开文件夹按钮
        self.pushButton1 = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton1.move(x, y+420)
        self.pushButton1.clicked.connect(On_button_clicked2) #按钮绑定槽函数

        #设置“项目文件夹”显示
        self.label9 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label9.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label9.resize(500, 20)#设置标签大小
        self.label9.setText("请选择导出的T++项目文件夹“data”")
        self.label9.move(x+150, y+425)   


        #设置“输出文件夹”标签
        self.label10 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label10.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label10.setText("输出文件夹")
        self.label10.move(x, y+480)  

        #设置输出文件夹按钮
        self.pushButton2 = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton2.move(x, y+520)
        self.pushButton2.clicked.connect(On_button_clicked3) #按钮绑定槽函数

        #设置“输出文件夹”显示
        self.label11 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label11.resize(500, 20)
        self.label11.setText("请选择翻译文件存储文件夹，不要与项目文件夹相同")
        self.label11.move(x+150, y+525)    



        #设置“开始翻译”的按钮
        self.primaryButton1 = PrimaryPushButton('开始翻译', self, FIF.UPDATE)
        self.primaryButton1.move(x+280, y+600)
        self.primaryButton1.clicked.connect(On_button_clicked6) #按钮绑定槽函数



        #设置“已花费”标签
        self.label22 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label22.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label22.setText("已花费")
        self.label22.move(x, y+640)
        self.label22.hide()  #先隐藏控件

        #设置“已花费金额”具体标签
        self.label23 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label23.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label23.resize(500, 20)#设置标签大小
        self.label23.setText("0＄")
        self.label23.move(x+60, y+640)
        self.label23.hide()  #先隐藏控件

        #设置翻译进度条控件
        self.progressBar2 = QProgressBar(self)
        self.progressBar2.setMinimum(0)
        self.progressBar2.setMaximum(100)
        self.progressBar2.setValue(0)
        self.progressBar2.setFixedSize(700, 30)
        self.progressBar2.move(x, y+670)
        self.progressBar2.setStyleSheet("QProgressBar::chunk { text-align: center; } QProgressBar { text-align: left; }")#使用setStyleSheet()方法设置了进度条块的文本居中对齐，并且设置了进度条的文本居左对齐
        self.progressBar2.setFormat("已翻译: %p%")
        self.progressBar2.hide()  #先隐藏控件

    def onCheckedChanged(self, isChecked: bool):
        if isChecked :
            self.SwitchButton1.setText("已开启")
            CreateWarningInfoBar("第一次使用语义检查功能，会下载500mb左右的模型，请注意网络环境与电脑存储空间！")
        else :
            self.SwitchButton1.setText("已关闭")


class Widget17(QFrame):#自定义的widget内容界面

    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用


        #设置各个控件-----------------------------------------------------------------------------------------


        #设置基础参数
        x = 60
        y = -230


        #设置“项目文件夹”标签
        self.label1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label1.setText("项目文件夹")
        self.label1.move(x, y+380)

        #设置打开文件夹按钮
        self.pushButton2 = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton2.move(x, y+420)
        self.pushButton2.clicked.connect(On_button_clicked2) #按钮绑定槽函数

        #设置“项目文件夹”显示
        self.label3 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label3.resize(500, 20)#设置标签大小
        self.label3.setText("请选择已翻译的T++项目文件夹“data”")
        self.label3.move(x+150, y+425)   


        #设置“输出文件夹”标签
        self.label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label4.setText("输出文件夹")
        self.label4.move(x, y+480)  

        #设置输出文件夹按钮
        self.pushButton5 = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton5.move(x, y+520)
        self.pushButton5.clicked.connect(On_button_clicked3) #按钮绑定槽函数

        #设置“输出文件夹”显示
        self.label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label6.resize(500, 20)
        self.label6.setText("请选择检查重翻存储文件夹，不要与原文件夹相同")
        self.label6.move(x+150, y+525)    


        #设置“开始检查”的按钮
        self.primaryButton7 = PrimaryPushButton('开始检查T++项目', self, FIF.UPDATE)
        self.primaryButton7.move(x+230, y+600)
        self.primaryButton7.clicked.connect(self.onChecked1) #按钮绑定槽函数




        #设置基础参数
        x1 = 60
        y1 = 100

        #设置“文件位置”标签
        self.label8 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label8.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label8.setText("文件位置")
        self.label8.move(x1, y1+380)

        #设置打开文件按钮
        self.pushButton1 = PushButton('选择文件', self, FIF.DOCUMENT)
        self.pushButton1.move(x1, y1+420)
        self.pushButton1.clicked.connect(On_button_clicked1) #按钮绑定槽函数

        #设置“文件位置”显示
        self.label9 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label9.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label9.resize(500, 20)#设置标签大小
        self.label9.setText("请选择需要已经翻译好的json文件")
        self.label9.move(x1+150, y1+425)   


        #设置“输出文件夹”标签
        self.label10 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label10.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label10.setText("输出文件夹")
        self.label10.move(x1, y1+480)  

        #设置输出文件夹按钮
        self.pushButton2 = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton2.move(x1, y1+520)
        self.pushButton2.clicked.connect(On_button_clicked3) #按钮绑定槽函数

        #设置“输出文件夹”显示
        self.label11 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label11.resize(500, 20)
        self.label11.setText("请选择检查重翻文件存储文件夹")
        self.label11.move(x1+150, y1+525)    



        #设置“开始检查”的按钮
        self.primaryButton1 = PrimaryPushButton('开始检查Mtool项目', self, FIF.UPDATE)
        self.primaryButton1.move(x1+230, y1+600)
        self.primaryButton1.clicked.connect(self.onChecked2) #按钮绑定槽函数

        #设置“已花费”标签
        self.label12 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label12.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label12.setText("已花费")
        self.label12.move(x1, y1+640)
        self.label12.hide()  #先隐藏控件

        #设置“已花费金额”具体标签
        self.label13 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label13.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label13.resize(500, 20)#设置标签大小
        self.label13.setText("0＄")
        self.label13.move(x1+60, y1+640)
        self.label13.hide()  #先隐藏控件

        #设置翻译进度条控件
        self.progressBar = QProgressBar(self)
        self.progressBar.setMinimum(0)
        self.progressBar.setMaximum(100)
        self.progressBar.setValue(0)
        self.progressBar.setFixedSize(700, 30)
        self.progressBar.move(x1, y1+670)
        self.progressBar.setStyleSheet("QProgressBar::chunk { text-align: center; } QProgressBar { text-align: left; }")#使用setStyleSheet()方法设置了进度条块的文本居中对齐，并且设置了进度条的文本居左对齐
        self.progressBar.setFormat("已翻译: %p%")
        self.progressBar.hide()  #先隐藏控件


    def onChecked1(self):
        global Running_status,money_used,Translation_Progress

        if Running_status == 0:
            
            Inspection_results = Config(2)   #读取配置信息，设置系统参数，并进行检查

            if Inspection_results == 0 :  #配置没有完全填写
                CreateErrorInfoBar("请正确填入配置信息,不要留空")
                Running_status = 0  #修改运行状态

            elif Inspection_results == 1 :  #账号类型和模型类型组合错误
                print("\033[1;31mError:\033[0m 请正确选择账号类型以及模型类型")
                Ui_signal.update_signal.emit("Wrong type selection")

            else :  
                #清空花销与进度，更新UI
                money_used = 0
                Translation_Progress = 0 

                Running_status = 5  #修改运行状态
                on_update_signal("Update_ui")
                OnButtonClicked("正在语义检查中" , "客官请耐心等待哦~~")

                #显示隐藏控件
                Window.Interface17.progressBar.show() 
                Window.Interface17.label12.show()
                Window.Interface17.label13.show() 


                #创建子线程
                thread = My_Thread()
                thread.start()



        elif Running_status == 1 or 2 or 3 or 4 or 5:
            CreateWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")

    def onChecked2(self):
        global Running_status,money_used,Translation_Progress

        if Running_status == 0:
            
            Inspection_results = Config(1)   #读取配置信息，设置系统参数，并进行检查

            if Inspection_results == 0 :  #配置没有完全填写
                CreateErrorInfoBar("请正确填入配置信息,不要留空")
                Running_status = 0  #修改运行状态

            elif Inspection_results == 1 :  #账号类型和模型类型组合错误
                print("\033[1;31mError:\033[0m 请正确选择账号类型以及模型类型")
                Ui_signal.update_signal.emit("Wrong type selection")

            else :  
                #清空花销与进度，更新UI
                money_used = 0
                Translation_Progress = 0 

                Running_status = 4  #修改运行状态
                on_update_signal("Update_ui")
                OnButtonClicked("正在语义检查中" , "客官请耐心等待哦~~")

                #显示隐藏控件
                Window.Interface17.progressBar.show() 
                Window.Interface17.label12.show()
                Window.Interface17.label13.show() 


                #创建子线程
                thread = My_Thread()
                thread.start()



        elif Running_status == 1 or 2 or 3 or 4 or 5:
            CreateWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")


class Widget18(QFrame):#自定义的widget内容界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用


        #设置基础坐标参数
        x=60
        y=340




        #设置“启用实时参数”标签
        self.label0 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label0.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label0.setText("启用调教功能")
        self.label0.move(x, y-140)

        #设置“启用实时参数”开关
        self.checkBox = CheckBox('实时设置AI参数', self)
        self.checkBox.move(x, y-100)
        self.checkBox.stateChanged.connect(self.checkBoxChanged)

        #设置官方文档说明链接按钮
        self.pushButton1 = PushButton('官方文档说明', self)
        self.pushButton1.move(x+500, y-100)
        self.pushButton1.clicked.connect(lambda: QDesktopServices.openUrl(QUrl('https://platform.openai.com/docs/api-reference/chat/create')))



        #设置“温度”标签
        self.label1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label1.setText("温度")
        self.label1.move(x, y)

        #设置“温度”副标签
        self.label11 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        self.label11.setText("官方默认值为1")
        self.label11.move(x+100, y+8)

        #设置“温度”滑动条
        self.slider1 = Slider(Qt.Horizontal, self)
        self.slider1.setFixedWidth(200)
        self.slider1.move(x, y+40)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label2 = QLabel(str(self.slider1.value()), self)
        self.label2.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.label2.move(x+210, y+44)
        self.slider1.valueChanged.connect(lambda value: self.label2.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label2显示正确的值
        self.slider1.setMinimum(0)
        self.slider1.setMaximum(20)
        self.slider1.setValue(0)

        #设置“温度”说明文档
        self.label3 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label3.setText("Temperature：控制结果的随机性。\n如果希望结果更有创意可以尝试 0.9 \n或者希望有固定结果可以尝试 0.0 \n官方建议不要与Top_p一同改变 ")
        self.label3.move(x, y+90)





        #设置“top_p”标签
        self.label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label4.setText("概率阈值")
        self.label4.move(x+400, y)

        #设置“top_p”副标签
        self.label41 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label41.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        self.label41.setText("官方默认值为1")
        self.label41.move(x+500, y+8)


        #设置“top_p”滑动条
        self.slider2 = Slider(Qt.Horizontal, self)
        self.slider2.setFixedWidth(200)
        self.slider2.move(x+400, y+40)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label5 = QLabel(str(self.slider2.value()), self)
        self.label5.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label5.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.label5.move(x+610, y+44)
        self.slider2.valueChanged.connect(lambda value: self.label5.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放在后面是为了让上面的label5显示正确的值和格式
        self.slider2.setMinimum(0)
        self.slider2.setMaximum(10)
        self.slider2.setValue(10)


        #设置“top_p”说明文档
        self.label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label6.setText("Top_p：用于控制生成文本的多样性。\n与Temperature的作用相同 \n如果希望结果更加多样可以尝试 0.9 \n或者希望有固定结果可以尝试 0.0\n官方建议不要与Temperature一同改变 ")
        self.label6.move(x+400, y+90)





        #设置“presence_penalty”标签
        self.label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label7.setText("主题惩罚")
        self.label7.move(x, y+200)

        #设置“presence_penalty”副标签
        self.label41 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label41.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        self.label41.setText("官方默认值为0")
        self.label41.move(x+100, y+208)


        #设置“presence_penalty”滑动条
        self.slider3 = Slider(Qt.Horizontal, self)
        self.slider3.setFixedWidth(200)
        self.slider3.move(x, y+240)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label8 = QLabel(str(self.slider3.value()), self)
        self.label8.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label8.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.label8.move(x+210, y+244)
        self.slider3.valueChanged.connect(lambda value: self.label8.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label8显示正确的值和格式
        self.slider3.setMinimum(-20)
        self.slider3.setMaximum(20)
        self.slider3.setValue(5)

        #设置“presence_penalty”说明文档
        self.label1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label1.setText("Presence_penalty：用于控制主题的重复度\n会根据到目前为止已经出现在文本中的语句\n正值是惩罚生成的新内容从而增加AI模型\n谈论新主题内容的可能性")
        self.label1.move(x, y+294)




        #设置“frequency_penalty”标签
        self.label9 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label9.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label9.setText("频率惩罚")
        self.label9.move(x+400, y+200)

        #设置“presence_penalty”副标签
        self.label91 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label91.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        self.label91.setText("官方默认值为0")
        self.label91.move(x+500, y+208)

        #设置“frequency_penalty”滑动条
        self.slider4 = Slider(Qt.Horizontal, self)
        self.slider4.setFixedWidth(200)
        self.slider4.move(x+400, y+240)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label10 = QLabel(str(self.slider4.value()), self)
        self.label10.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.label10.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label10.move(x+610, y+244)
        self.slider4.valueChanged.connect(lambda value: self.label10.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label10显示正确的值和格式
        self.slider4.setMinimum(-20)
        self.slider4.setMaximum(20)
        self.slider4.setValue(0)

        #设置“frequency_penalty”说明文档
        self.label11 = QLabel(parent=self, flags=Qt.WindowFlags())
        self.label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px;  color: black")
        self.label11.setText("Frequency_penalty：会根据新词在文本中\n的现有频率，负值进行奖励，正值进行惩罚\n以便增加或降低逐字重复同一行的可能性")
        self.label11.move(x+400, y+294)

    
    # 勾选事件
    def checkBoxChanged(self, isChecked: bool):
        if isChecked :
            CreateSuccessInfoBar("已启用实时调教功能")


class AvatarWidget(NavigationWidget):#自定义的头像导航项
    """ Avatar widget """

    def __init__(self, parent=None):
        super().__init__(isSelectable=False, parent=parent)
        self.avatar = QImage('resource/Avatar.png').scaled(
            24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation)

    def paintEvent(self, e):
        painter = QPainter(self)
        painter.setRenderHints(
            QPainter.SmoothPixmapTransform | QPainter.Antialiasing)

        painter.setPen(Qt.NoPen)

        if self.isPressed:
            painter.setOpacity(0.7)

        # draw background
        if self.isEnter:
            c = 255 if isDarkTheme() else 0
            painter.setBrush(QColor(c, c, c, 10))
            painter.drawRoundedRect(self.rect(), 5, 5)

        # draw avatar
        painter.setBrush(QBrush(self.avatar))
        painter.translate(8, 6)
        painter.drawEllipse(0, 0, 24, 24)
        painter.translate(-8, -6)

        if not self.isCompacted:
            painter.setPen(Qt.white if isDarkTheme() else Qt.black)
            font = QFont('Segoe UI')
            font.setPixelSize(14)
            painter.setFont(font)
            painter.drawText(QRect(44, 0, 255, 36), Qt.AlignVCenter, 'NEKOparapa')


class CustomTitleBar(TitleBar): #自定义的标题栏
    """ Title bar with icon and title """

    def __init__(self, parent):
        super().__init__(parent)
        # add window icon
        self.iconLabel = QLabel(self) #创建标签
        self.iconLabel.setFixedSize(18, 18) #设置标签大小
        self.hBoxLayout.insertSpacing(0, 10) #设置布局的间距
        self.hBoxLayout.insertWidget(1, self.iconLabel, 0, Qt.AlignLeft | Qt.AlignBottom) #将标签添加到布局中
        self.window().windowIconChanged.connect(self.setIcon) #窗口图标改变时，调用setIcon函数

        # add title label
        self.titleLabel = QLabel(self) #创建标签
        self.hBoxLayout.insertWidget(2, self.titleLabel, 0, Qt.AlignLeft | Qt.AlignBottom) #将标签添加到布局中
        self.titleLabel.setObjectName('titleLabel') #设置对象名
        self.window().windowTitleChanged.connect(self.setTitle) #窗口标题改变时，调用setTitle函数

    def setTitle(self, title): #设置标题
        self.titleLabel.setText(title) #设置标签的文本
        self.titleLabel.adjustSize() #调整标签的大小

    def setIcon(self, icon): #设置图标
        self.iconLabel.setPixmap(QIcon(icon).pixmap(18, 18)) #设置图标


class window(FramelessWindow): #自定义的窗口

    def __init__(self):
        super().__init__()
        # use dark theme mode
        setTheme(Theme.LIGHT) #设置主题

        self.hBoxLayout = QHBoxLayout(self) #设置布局为水平布局

        self.setTitleBar(CustomTitleBar(self)) #设置标题栏，传入参数为自定义的标题栏
        self.stackWidget = QStackedWidget(self) #创建堆栈父2窗口
        self.navigationInterface = NavigationInterface(
            self, showMenuButton=True, showReturnButton=True) #创建父3导航栏


        # create sub interface
        self.Interface11 = Widget11('Interface11', self)     #创建子界面Interface(搜索界面)，传入参数为对象名和parent
        self.Interface12 = Widget12('Interface12', self)     #创建子界面Interface(文件夹界面)，传入参数为对象名和parent
        self.Interface15 = Widget15('Interface15', self)      #创建子界面Interface(音乐界面) ，传入参数为对象名和parent
        self.Interface16 = Widget16('Interface16', self)        #创建子界面Interface(视频界面)，传入参数为对象名和parent
        self.Interface17 = Widget17('Interface17', self) 
        self.Interface18 = Widget18('Interface18', self) 


        self.stackWidget.addWidget(self.Interface11)  #将子界面添加到父2堆栈窗口中
        self.stackWidget.addWidget(self.Interface12)
        self.stackWidget.addWidget(self.Interface15)
        self.stackWidget.addWidget(self.Interface16)
        self.stackWidget.addWidget(self.Interface17)
        self.stackWidget.addWidget(self.Interface18)


        self.initLayout() #调用初始化布局函数 

        self.initNavigation()   #调用初始化导航栏函数

        self.initWindow()  #调用初始化窗口函数



    #初始化布局的函数
    def initLayout(self):   
        self.hBoxLayout.setSpacing(0)                   #设置水平布局的间距
        self.hBoxLayout.setContentsMargins(0, 0, 0, 0)   #设置水平布局的边距
        self.hBoxLayout.addWidget(self.navigationInterface)    #将导航栏添加到布局中
        self.hBoxLayout.addWidget(self.stackWidget)            #将堆栈窗口添加到布局中
        self.hBoxLayout.setStretchFactor(self.stackWidget, 1) #设置堆栈窗口的拉伸因子

        self.titleBar.raise_() #将标题栏置于顶层
        self.navigationInterface.displayModeChanged.connect(self.titleBar.raise_) #导航栏的显示模式改变时，将标题栏置于顶层

    #初始化导航栏的函数
    def initNavigation(self): #详细介绍：https://pyqt-fluent-widgets.readthedocs.io/zh_CN/latest/navigation.html


        self.navigationInterface.addItem(  #addItem函数是导航栏的函数，用于添加导航项
            routeKey=self.Interface11.objectName(), #设置路由键,路由键是导航项的唯一标识符,用于切换导航项,这里设置为子界面的对象名
            icon=FIF.FEEDBACK, #设置左侧图标
            text='官方账号',  #设置显示文本
            onClick=lambda: self.switchTo(self.Interface11) #设置点击事件
        )   #添加导航项，传入参数：路由键，图标，文本，点击事件


        #添加国内代理导航项
        self.navigationInterface.addItem(
            routeKey=self.Interface12.objectName(),
            icon=FIF.FEEDBACK,
            text='代理账号',
            onClick=lambda: self.switchTo(self.Interface12),
            #position=NavigationItemPosition.SCROLL #设置导航项的位置
            ) 
        
        self.navigationInterface.addSeparator() #添加分隔符

        self.navigationInterface.addItem(
            routeKey=self.Interface15.objectName(),
            icon=FIF.BOOK_SHELF,
            text='Mtool项目',
            onClick=lambda: self.switchTo(self.Interface15)
        )  #添加导航项
        self.navigationInterface.addItem(
            routeKey=self.Interface16.objectName(),
            icon=FIF.BOOK_SHELF,
            text='Translator++项目',
            onClick=lambda: self.switchTo(self.Interface16)
        ) #添加导航项

        self.navigationInterface.addSeparator() #添加分隔符

        
        #添加词义检查导航项
        self.navigationInterface.addItem(
            routeKey=self.Interface17.objectName(),
            icon=FIF.HIGHTLIGHT,
            text='语义检查',
            onClick=lambda: self.switchTo(self.Interface17),
            position=NavigationItemPosition.SCROLL
            ) 


        #添加测试导航项
        self.navigationInterface.addItem(
            routeKey=self.Interface18.objectName(),
            icon=FIF.ALBUM,
            text='实时调教',
            onClick=lambda: self.switchTo(self.Interface18),
            position=NavigationItemPosition.SCROLL
            ) 




       # 添加头像导航项
        self.navigationInterface.addWidget(
            routeKey='avatar',
            widget=AvatarWidget(),
            onClick=self.showMessageBox,
            position=NavigationItemPosition.BOTTOM
        )


        #!IMPORTANT: don't forget to set the default route key
        self.navigationInterface.setDefaultRouteKey(self.Interface11.objectName()) #设置默认的路由键,不起作用
        

        # set the maximum width
        # self.navigationInterface.setExpandWidth(300)

        self.stackWidget.currentChanged.connect(self.onCurrentInterfaceChanged) #堆栈窗口的当前窗口改变时，调用onCurrentInterfaceChanged函数
        self.stackWidget.setCurrentIndex(1) #设置堆栈窗口的当前窗口为1

    #头像导航项的函数调用的函数
    def showMessageBox(self):
        url = QUrl('https://github.com/NEKOparapa/AiNiee-chatgpt')
        QDesktopServices.openUrl(url)

    #初始化父窗口的函数
    def initWindow(self): 
        self.resize(850, 820) #设置窗口的大小
        #self.setWindowIcon(QIcon('resource/logo.png')) #设置窗口的图标
        self.setWindowTitle(Software_Version) #设置窗口的标题
        self.titleBar.setAttribute(Qt.WA_StyledBackground) #设置标题栏的属性

        # 移动到屏幕中央
        desktop = QApplication.desktop().availableGeometry() #获取桌面的可用几何
        w, h = desktop.width(), desktop.height() #获取桌面的宽度和高度
        self.move(w//2 - self.width()//2, h//2 - self.height()//2) #将窗口移动到桌面的中心


        #根据主题设置设置样式表的函数
        color = 'dark' if isDarkTheme() else 'light' #如果是暗色主题，则color为dark，否则为light
        with open(f'resource/{color}/demo.qss', encoding='utf-8') as f: #打开样式表
            self.setStyleSheet(f.read()) #设置样式表

    #切换到某个窗口的函数
    def switchTo(self, widget): 
        self.stackWidget.setCurrentWidget(widget) #设置堆栈窗口的当前窗口为widget

    #堆栈窗口的当前窗口改变时，调用的函数
    def onCurrentInterfaceChanged(self, index):    
        widget = self.stackWidget.widget(index) #获取堆栈窗口的当前窗口
        self.navigationInterface.setCurrentItem(widget.objectName()) #设置导航栏的当前项为widget的对象名

    #重写鼠标按下事件
    def resizeEvent(self, e): 
        self.titleBar.move(46, 0) #将标题栏移动到(46, 0)
        self.titleBar.resize(self.width()-46, self.titleBar.height()) #设置标题栏的大小

    #窗口关闭函数，放在最后面，解决界面空白与窗口退出后子线程还在运行的问题
    def closeEvent(self, event):
        title = '确定是否退出程序?'
        content = """如果正在进行翻译任务，当前任务会停止,并备份已经翻译的内容。"""
        w = Dialog(title, content, self)

        if w.exec() :
            print("[INFO] 主窗口已经退出！")
            global Running_status
            Running_status = 10
            event.accept()
        else:
            event.ignore()


if __name__ == '__main__':

    #开启子进程支持
    multiprocessing.freeze_support() 

    # 启用了高 DPI 缩放
    QApplication.setHighDpiScaleFactorRoundingPolicy(
        Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)


    # 创建子线程通信的信号
    Ui_signal = UI_signal() #创建子线程类，并创建新信号
    Ui_signal.update_signal.connect(on_update_signal)  #创建信号与槽函数的绑定


    #创建了一个 QApplication 对象
    app = QApplication(sys.argv)
    #创建窗口对象
    Window = window()
    
    #窗口对象显示
    Window.show()

    #读取配置文件
    Read_Write_Config("read") 


    #进入事件循环，等待用户操作
    sys.exit(app.exec_())



